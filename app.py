# --- STOCK REGISTRO ENDPOINT ---
# (Movido después de la inicialización de Flask y Firebase)

from flask import Flask, request, jsonify, redirect, render_template
import firebase_admin
from firebase_admin import credentials, auth, db as firebase_db
from google.cloud import firestore
import os
import json

print('Iniciando Flask...')

# Inicializar Firebase con credenciales desde variable de entorno o archivo local
if os.environ.get('FIREBASE_CREDENTIALS'):
    # En producción (Railway): leer desde variable de entorno
    print('Usando credenciales de Firebase desde variable de entorno')
    firebase_creds = json.loads(os.environ.get('FIREBASE_CREDENTIALS'))
    cred = credentials.Certificate(firebase_creds)
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://supratechweb-default-rtdb.firebaseio.com/'
        })
    db = firestore.Client.from_service_account_info(firebase_creds)
else:
    # En desarrollo local: leer desde archivo
    print('Usando credenciales de Firebase desde archivo local')
    cred = credentials.Certificate('supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json')
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://supratechweb-default-rtdb.firebaseio.com/'
        })
    db = firestore.Client.from_service_account_json('supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json')

app = Flask(__name__)

# Función auxiliar para obtener credenciales de Google API
def get_google_credentials(scopes=None):
    """Retorna credenciales de Google API desde variable de entorno o archivo local"""
    from google.oauth2 import service_account
    if scopes is None:
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    else:
        SCOPES = scopes

    if os.environ.get('FIREBASE_CREDENTIALS'):
        # En producción: usar credenciales desde variable de entorno
        firebase_creds = json.loads(os.environ.get('FIREBASE_CREDENTIALS'))
        return service_account.Credentials.from_service_account_info(firebase_creds, scopes=SCOPES)
    else:
        # En desarrollo: usar archivo local
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        return service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# --- ENDPOINTS PARA BDMARCAS ---
@app.route('/api/bdmarcas_campos', methods=['GET'])
def bdmarcas_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDMarcas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/api/bdmarcas_registro', methods=['POST'])
def bdmarcas_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import datetime, re, traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDMarcas')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDMarcas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=rango_col,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            # Extraer columna y fila inicial correctamente (ej: A2:A, B5:B, etc.)
            import re
            partes = rango.split(':')[0]
            col_match = re.match(r"^([A-Z]+)([0-9]+)?$", partes)
            if col_match:
                col = col_match.group(1)
                fila_inicial = int(col_match.group(2)) if col_match.group(2) else 1
            else:
                col = partes
                fila_inicial = 1
            values = [[data.get(campo, '')]]
            rango_celda = f"{nombre_hoja}!{col}{fila_inicial}:{col}{fila_inicial}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bdmarcas_bulk', methods=['POST'])
def bdmarcas_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback, sys, re, io, csv, openpyxl
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDMarcas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDMarcas')
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        rows = []
        if filename.endswith('.csv'):
            try:
                stream = io.StringIO(file.stream.read().decode('utf-8'))
            except Exception:
                stream = io.StringIO(file.stream.read().decode('latin-1'))
            reader = csv.DictReader(stream)
            if reader.fieldnames is None:
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=rango_col,
                body={}
            ).execute()
        import re
        for campo, rango in ubicaciones.items():
            # Extraer columna y fila inicial correctamente (ej: G7:G)
            match = re.match(r"([A-Z]+)(\d+):", rango)
            if match:
                col = match.group(1)
                fila_inicial = int(match.group(2))
            else:
                # fallback: solo columna, empieza en 1
                col = rango.split(':')[0]
                fila_inicial = 1
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila_inicial + len(values) - 1
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=f"{nombre_hoja}!{col}{fila_inicial}:{col}{fila_final}",
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Endpoint de login para recibir el token de Firebase y devolver info básica del usuario
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    id_token = data.get('idToken')
    if not id_token:
        return jsonify({'error': 'No token provided'}), 401
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        email = decoded_token.get('email', '')
        # Puedes personalizar la consulta a Firestore para obtener más info del usuario
        user_ref = db.collection('users').document(uid)
        user_doc = user_ref.get()
        user_data = user_doc.to_dict() if user_doc.exists else {}
        return jsonify({'status': 'ok', 'uid': uid, 'email': email, 'rol': user_data.get('rol', 'usuario')})
    except Exception as e:
        return jsonify({'error': 'Invalid token', 'details': str(e)}), 401


@app.route('/api/baseplus_campos', methods=['GET'])
def baseplus_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BASEPLUS')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/')
def home():
    return render_template('demo.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/panel')
def panel():
    # Ejemplo básico: datos de usuario de prueba
    user = {
        'nombre': 'Usuario',
        'email': 'usuario@email.com',
        'rol': 'usuario'
    }
    esquema_raw = 'N/A'
    uid = 'uid_de_ejemplo'
    return render_template('panel.html', user=user, esquema_raw=esquema_raw, uid=uid)

@app.route('/api/get_spreadsheet', methods=['POST'])
def get_spreadsheet():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        user_ref = db.collection('users').document(uid)
        user_doc = user_ref.get()
        if user_doc.exists:
            user_data = user_doc.to_dict()
            spreadsheet_id = user_data.get('spreadsheetId')
            if spreadsheet_id:
                return jsonify({'spreadsheetId': spreadsheet_id})
            else:
                return jsonify({'error': 'No Software asignado aún'}), 404
        else:
            return jsonify({'error': 'Usuario no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': 'Invalid token', 'details': str(e)}), 401




@app.route('/panel_opciones')
def panel_opciones():
    return render_template('panel_opciones.html')

# Nueva ruta para BASEPLUS
@app.route('/baseplus')
def baseplus():
    return render_template('BASEPLUS.html')


@app.route('/Llenar_BDS')
def llenar_bds():
    return render_template('Llenar_BDS.html')

# Nueva ruta para BD MARCAS
@app.route('/BDMARCAS')
def bdmarcas():
    return render_template('BDMARCAS.html')

# Ruta para HISTORICO
@app.route('/historico')
def historico():
    return render_template('HISTORICO.html')

@app.route('/api/userinfo', methods=['GET'])
def userinfo():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        user_ref = db.collection('users').document(uid)
        user_doc = user_ref.get()
        if user_doc.exists:
            user_data = user_doc.to_dict()
            user_data['uid'] = uid
            return jsonify(user_data)
        else:
            return jsonify({'error': 'Usuario no encontrado'}), 404
    except Exception as e:
        return jsonify({'error': 'Invalid token', 'details': str(e)}), 401


# Endpoint para registrar datos en BASEPLUS (spreadsheets personalizados por usuario)

@app.route('/api/baseplus_registro', methods=['POST'])
def baseplus_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        print('[BASEPLUS_REGISTRO] UID:', uid)
        print('[BASEPLUS_REGISTRO] Data recibida:', data)
        # Buscar la URL del spreadsheet en la colección Areas, campo BASEPLUS
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            print('[BASEPLUS_REGISTRO] No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BASEPLUS')
        print('[BASEPLUS_REGISTRO] Spreadsheet URL:', spreadsheet_url)
        if not spreadsheet_url:
            print('[BASEPLUS_REGISTRO] No hay spreadsheet asignado para BASEPLUS')
            return jsonify({'error': 'No hay spreadsheet asignado para BASEPLUS'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            print('[BASEPLUS_REGISTRO] URL de spreadsheet inválida')
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        print('[BASEPLUS_REGISTRO] Spreadsheet ID:', spreadsheet_id)

        # Leer hoja y ubicaciones desde /Areas/{uid}/Hojas/BASEPLUS
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BASEPLUS')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            print('[BASEPLUS_REGISTRO] No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        print('[BASEPLUS_REGISTRO] Hoja data:', hoja_data)
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        # Quitar el campo 'Hoja' para solo dejar los campos de ubicaciones
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[BASEPLUS_REGISTRO] Ubicaciones:', ubicaciones)
        # Construir los valores a insertar
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import datetime
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # --- NUEVA LÓGICA: Limpiar todos los rangos configurados antes de registrar y escribir SIEMPRE en la primera fila del rango configurado ---
        print(f'[BASEPLUS_REGISTRO] Limpiando todos los rangos configurados antes de registrar...')
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            print(f'[BASEPLUS_REGISTRO] Limpiando rango {rango_col}')
            try:
                sheet.values().clear(
                    spreadsheetId=spreadsheet_id,
                    range=rango_col,
                    body={}
                ).execute()
            except Exception as e:
                print(f'[BASEPLUS_REGISTRO] Error limpiando rango {rango_col}:', str(e))
                print(traceback.format_exc())

        # Escribir cada campo en su rango configurado individualmente
        for campo, rango in ubicaciones.items():
            valor = data.get(campo, '')
            rango_celda = f"{nombre_hoja}!{rango.split(':')[0]}"
            print(f'[BASEPLUS_REGISTRO] Insertando {campo} en {rango_celda}:', valor)
            try:
                sheet.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=rango_celda,
                    valueInputOption='USER_ENTERED',
                    body={'values': [[valor]]}
                ).execute()
            except Exception as e:
                print(f'[BASEPLUS_REGISTRO] Error insertando {campo} en {rango_celda}:', str(e))
                print(traceback.format_exc())
                return jsonify({'error': f'Error insertando {campo} en {rango_celda}', 'details': str(e)}), 500
        print('[BASEPLUS_REGISTRO] Registro exitoso')
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[BASEPLUS_REGISTRO] Error general:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en BASEPLUS', 'details': str(e)}), 500


# --- BASEPLUS BULK ENDPOINT ---
from werkzeug.utils import secure_filename
import os
import csv
import io
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

@app.route('/api/baseplus_bulk', methods=['POST'])
def baseplus_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    import sys
    def log(msg, *args):
        print('[BULK]', msg, *args)
        sys.stdout.flush()
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        log('UID:', uid)
        db = firestore.Client.from_service_account_json('supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json')
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BASEPLUS')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            log('No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            log('No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BASEPLUS')
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            log('URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            log('No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError as e:
                log('Fallo utf-8, intentando latin-1:', str(e))
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            log(f'Limpiando rango {rango_col}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=rango_col,
                body={}
            ).execute()
        import re
        for campo, rango in ubicaciones.items():
            # Extraer columna y fila inicial correctamente (ej: G7:G)
            match = re.match(r"([A-Z]+)(\d+):", rango)
            if match:
                col = match.group(1)
                fila_inicial = int(match.group(2))
            else:
                # fallback: solo columna, empieza en 1
                col = rango.split(':')[0]
                fila_inicial = 1
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila_inicial + len(values) - 1
            log(f'Escribiendo campo {campo} en {col}{fila_inicial}:{col}{fila_final} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{fila_inicial}:{col}{fila_final}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Endpoint para registrar compras en el spreadsheet del usuario
from google.oauth2 import service_account
from googleapiclient.discovery import build
import datetime

@app.route('/api/registrar_compra', methods=['POST'])
def registrar_compra():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        # Obtener datos de la compra
        data = request.get_json()
        producto = data.get('producto')
        cantidad = data.get('cantidad')
        precio = data.get('precio')
        if not producto or not cantidad or not precio:
            return jsonify({'error': 'Datos incompletos'}), 400
        # Buscar la URL del spreadsheet en la colección Areas, campo Compras
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Compras')
        if not spreadsheet_url:
            return jsonify({'error': 'No hay spreadsheet asignado para Compras'}), 404
        # Extraer el ID del spreadsheet desde la URL
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)

        # Leer hoja y rango desde /Areas/{uid}/Hojas/Desgloce
        desgloce_ref = db.collection('Areas').document(uid).collection('Hojas').document('Desgloce')
        desgloce_doc = desgloce_ref.get()
        if not desgloce_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        desgloce_data = desgloce_doc.to_dict()
        nombre_hoja = desgloce_data.get('Hoja', 'Sheet1')
        rango = desgloce_data.get('Rango', 'A1')
        # El rango final será "{nombre_hoja}!{rango}", por ejemplo "Compras2!A2:B"
        rango_final = f"{nombre_hoja}!{rango}"

        # Autenticación con Google Sheets API
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        # Registrar la compra (agregar fila)
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # Solo enviar los campos que correspondan al rango
        # Si el rango es A2:B, solo enviar dos columnas: producto y cantidad
        columnas = rango.split(':')
        num_cols = 1
        if len(columnas) == 2:
            # Calcular número de columnas entre A y B, etc.
            def col_to_num(col):
                num = 0
                for c in col:
                    num = num * 26 + (ord(c.upper()) - ord('A') + 1)
                return num
            num_cols = col_to_num(columnas[1].rstrip('0123456789')) - col_to_num(columnas[0].rstrip('0123456789')) + 1
        # Preparar los valores según el número de columnas
        all_values = [fecha, producto, cantidad, precio]
        values = [all_values[:num_cols]]
        body = {'values': values}
        result = sheet.values().append(
            spreadsheetId=spreadsheet_id,
            range=rango_final,
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body=body
        ).execute()
        return jsonify({'status': 'ok', 'updatedRange': result.get('updates', {}).get('updatedRange', '')})
    except Exception as e:
        return jsonify({'error': 'Error al registrar la compra', 'details': str(e)}), 500

# Nueva ruta para STOCK
@app.route('/stock')
def stock():
    return render_template('Stock.html')

@app.route('/api/stock', methods=['GET'])
def stock_api():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('STOCK')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/api/stock_campos', methods=['GET'])
def stock_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Stock')  # Igual que en Firebase
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500


# --- STOCK BULK ENDPOINT ---
from werkzeug.utils import secure_filename
import os
import csv
import io
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

@app.route('/api/stock_bulk', methods=['POST'])
def stock_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        import sys
        def log(msg, *args):
            print('[STOCK_BULK]', msg, *args)
            sys.stdout.flush()
        log('UID:', uid)

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Stock')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            log('No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            log('No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Stock')
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            log('URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            log('No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError as e:
                log('Fallo utf-8, intentando latin-1:', str(e))
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        # Limpiar todos los rangos configurados antes de registrar y escribir SIEMPRE en la fila 2
        import re
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente (ej: A22:A, B5:B, etc.) y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            log(f'Limpiando rango {clear_range}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=clear_range,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            fila = 2  # Siempre iniciar en la fila 2
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila + len(values) - 1
            log(f'Escribiendo campo {campo} en {col}{fila}:{col}{fila_final} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila_final}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- STOCK REGISTRO ENDPOINT ---
@app.route('/api/stock_registro', methods=['POST'])
def stock_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        print('[STOCK_REGISTRO] UID:', uid)
        print('[STOCK_REGISTRO] Data recibida:', data)
        # Buscar la URL del spreadsheet en la colección Areas, campo Stock
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            print('[STOCK_REGISTRO] No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Stock')
        print('[STOCK_REGISTRO] Spreadsheet URL:', spreadsheet_url)
        if not spreadsheet_url:
            print('[STOCK_REGISTRO] No hay spreadsheet asignado para Stock')
            return jsonify({'error': 'No hay spreadsheet asignado para Stock'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            print('[STOCK_REGISTRO] URL de spreadsheet inválida')
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        print('[STOCK_REGISTRO] Spreadsheet ID:', spreadsheet_id)

        # Leer hoja y ubicaciones desde /Areas/{uid}/Hojas/Stock
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Stock')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            print('[STOCK_REGISTRO] No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        print('[STOCK_REGISTRO] Hoja data:', hoja_data)
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[STOCK_REGISTRO] Ubicaciones:', ubicaciones)
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import datetime
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'[STOCK_REGISTRO] Limpiando todos los rangos configurados antes de registrar...')
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            print(f'[STOCK_REGISTRO] Limpiando rango {rango_col}')
            try:
                sheet.values().clear(
                    spreadsheetId=spreadsheet_id,
                    range=rango_col,
                    body={}
                ).execute()
            except Exception as e:
                print(f'[STOCK_REGISTRO] Error limpiando rango {rango_col}:', str(e))
                print(traceback.format_exc())
        for campo, rango in ubicaciones.items():
            valor = data.get(campo, '')
            # Extraer columna (ej: A2:A, B5:B, etc.) y forzar fila 2
            import re
            partes = rango.split(':')[0]
            col_match = re.match(r"^([A-Z]+)([0-9]+)?$", partes)
            if col_match:
                col = col_match.group(1)
            else:
                col = partes
            fila_inicial = 2
            rango_celda = f"{nombre_hoja}!{col}{fila_inicial}:{col}{fila_inicial}"
            print(f'[STOCK_REGISTRO] Insertando {campo} en {rango_celda}:', valor)
            try:
                sheet.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=rango_celda,
                    valueInputOption='USER_ENTERED',
                    body={'values': [[valor]]}
                ).execute()
            except Exception as e:
                print(f'[STOCK_REGISTRO] Error insertando {campo} en {rango_celda}:', str(e))
                print(traceback.format_exc())
                return jsonify({'error': f'Error insertando {campo} en {rango_celda}', 'details': str(e)}), 500
        print('[STOCK_REGISTRO] Registro exitoso')
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[STOCK_REGISTRO] Error general:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Stock', 'details': str(e)}), 500

# Nueva ruta para VENTAS
@app.route('/ventas')
def ventas():
    return render_template('Ventas.html')

@app.route('/api/ventas_campos', methods=['GET'])
def ventas_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Ventas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

# --- VENTAS REGISTRO ENDPOINT ---
@app.route('/api/ventas_registro', methods=['POST'])
def ventas_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import datetime, re, traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        print('[VENTAS_REGISTRO] UID:', uid)
        print('[VENTAS_REGISTRO] Data recibida:', data)
        # Buscar la URL del spreadsheet en la colección Areas, campo Ventas
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Ventas')
        print('[VENTAS_REGISTRO] Spreadsheet URL:', spreadsheet_url)
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        print('[VENTAS_REGISTRO] Spreadsheet ID:', spreadsheet_id)

        # Leer hoja y ubicaciones desde /Areas/{uid}/Hojas/Ventas
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Ventas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        print('[VENTAS_REGISTRO] Hoja data:', hoja_data)
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[VENTAS_REGISTRO] Ubicaciones:', ubicaciones)
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import datetime
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'[VENTAS_REGISTRO] Limpiando todos los rangos configurados antes de registrar...')
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            print(f'[VENTAS_REGISTRO] Limpiando rango {clear_range}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=clear_range,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            valor = data.get(campo, '')
            if campo.lower() == 'fecha':
                valor = fecha
            values = [[valor]]
            rango_celda = f"{nombre_hoja}!{col}2:{col}2"
            print(f'[VENTAS_REGISTRO] Escribiendo campo {campo} en {rango_celda} con valor: {valor}')
            try:
                sheet.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=rango_celda,
                    valueInputOption='USER_ENTERED',
                    body={'values': values}
                ).execute()
            except Exception as e:
                print(f'[VENTAS_REGISTRO] Error insertando {campo} en {rango_celda}:', str(e))
                print(traceback.format_exc())
                return jsonify({'error': f'Error insertando {campo} en {rango_celda}', 'details': str(e)}), 500
        print('[VENTAS_REGISTRO] Registro exitoso')
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[VENTAS_REGISTRO] Error general:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Ventas', 'details': str(e)}), 500

# --- VENTAS BULK ENDPOINT ---
@app.route('/api/ventas_bulk', methods=['POST'])
def ventas_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        import sys
        def log(msg, *args):
            print('[VENTAS_BULK]', msg, *args)
            sys.stdout.flush()
        log('UID:', uid)

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Ventas')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            log('No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            log('No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Ventas')
        if not spreadsheet_url:
            log('No se encontró la URL del spreadsheet para Ventas')
            return jsonify({'error': 'No se encontró la URL del spreadsheet para Ventas. Configura el campo "Ventas" en Firestore.'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            log('URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            log('No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            import csv, io
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError as e:
                log('Fallo utf-8, intentando latin-1:', str(e))
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        # Limpiar todos los rangos configurados antes de registrar y escribir SIEMPRE en la fila 2
        import re
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente (ej: A22:A, B5:B, etc.) y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            log(f'Limpiando rango {clear_range}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=clear_range,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            # Extraer columna correctamente y forzar fila 2
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            fila = 2  # Siempre iniciar en la fila 2
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila + len(values) - 1
            log(f'Escribiendo campo {campo} en {col}{fila}:{col}{fila_final} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila_final}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# Nueva ruta para BLACKLIST
@app.route('/blacklist')
def blacklist():
    return render_template('Blacklist.html')

# Nueva ruta para PEDIDOS 9.0
@app.route('/pedidos')
def pedidos():
    return render_template('Pedidos.html')

# Rutas para submódulos de Pedidos 9.0
@app.route('/pedidos/calendario')
def pedidos_calendario():
    return render_template('Pedidos_Calendario.html')

@app.route('/pedidos/bd')
def pedidos_bd():
    return render_template('Pedidos_BD.html')

@app.route('/pedidos/bdqty')
def pedidos_bdqty():
    return render_template('Pedidos_BDQTY.html')

@app.route('/pedidos/clasificaciones')
def pedidos_clasificaciones():
    return render_template('Pedidos_Clasificaciones.html')

@app.route('/api/blacklist_campos', methods=['GET'])
def blacklist_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Blacklist')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

# --- BLACKLIST REGISTRO ENDPOINT ---
@app.route('/api/blacklist_registro', methods=['POST'])
def blacklist_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import datetime, re, traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        print('[BLACKLIST_REGISTRO] UID:', uid)
        print('[BLACKLIST_REGISTRO] Data recibida:', data)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Blacklist')
        print('[BLACKLIST_REGISTRO] Spreadsheet URL:', spreadsheet_url)
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        print('[BLACKLIST_REGISTRO] Spreadsheet ID:', spreadsheet_id)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Blacklist')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        print('[BLACKLIST_REGISTRO] Hoja data:', hoja_data)
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[BLACKLIST_REGISTRO] Ubicaciones:', ubicaciones)
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import datetime
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Encontrar la siguiente fila vacía usando 'Marca' como referencia
        primer_campo = 'Marca' if 'Marca' in ubicaciones else list(ubicaciones.keys())[0]
        primer_rango = ubicaciones[primer_campo]
        match = re.match(r"([A-Z]+)", primer_rango)
        if match:
            col = match.group(1)
        else:
            col = primer_rango.split(':')[0]
        
        rango_lectura = f"{nombre_hoja}!{col}2:{col}"
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=rango_lectura
        ).execute()
        valores_existentes = result.get('values', [])
        siguiente_fila = len(valores_existentes) + 2
        print(f'[BLACKLIST_REGISTRO] Siguiente fila disponible: {siguiente_fila}')
        
        # Escribir en la siguiente fila vacía
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            valor = data.get(campo, '')
            if campo.lower() == 'fecha':
                valor = fecha
            values = [[valor]]
            rango_celda = f"{nombre_hoja}!{col}{siguiente_fila}:{col}{siguiente_fila}"
            print(f'[BLACKLIST_REGISTRO] Escribiendo campo {campo} en {rango_celda} con valor: {valor}')
            try:
                sheet.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=rango_celda,
                    valueInputOption='USER_ENTERED',
                    body={'values': values}
                ).execute()
            except Exception as e:
                print(f'[BLACKLIST_REGISTRO] Error insertando {campo} en {rango_celda}:', str(e))
                print(traceback.format_exc())
                return jsonify({'error': f'Error insertando {campo} en {rango_celda}', 'details': str(e)}), 500
        print('[BLACKLIST_REGISTRO] Registro exitoso')
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[BLACKLIST_REGISTRO] Error general:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Blacklist', 'details': str(e)}), 500

# --- BLACKLIST BULK ENDPOINT ---
@app.route('/api/blacklist_bulk', methods=['POST'])
def blacklist_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        import sys
        def log(msg, *args):
            print('[BLACKLIST_BULK]', msg, *args)
            sys.stdout.flush()
        log('UID:', uid)

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Blacklist')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            log('No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            log('No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Blacklist')
        if not spreadsheet_url:
            log('No se encontró la URL del spreadsheet para Blacklist')
            return jsonify({'error': 'No se encontró la URL del spreadsheet para Blacklist. Configura el campo "Blacklist" en Firestore.'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            log('URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            log('No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            import csv, io
            try:
                content = file.stream.read().decode('utf-8-sig')  # utf-8-sig elimina BOM automáticamente
            except UnicodeDecodeError as e:
                log('Fallo utf-8, intentando latin-1:', str(e))
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        import re
        rows_preview = rows[:100]
        
        # Encontrar la siguiente fila vacía usando 'Marca' como referencia
        primer_campo = 'Marca' if 'Marca' in ubicaciones else list(ubicaciones.keys())[0]
        primer_rango = ubicaciones[primer_campo]
        match = re.match(r"([A-Z]+)", primer_rango)
        if match:
            col = match.group(1)
        else:
            col = primer_rango.split(':')[0]
        
        rango_lectura = f"{nombre_hoja}!{col}2:{col}"
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=rango_lectura
        ).execute()
        valores_existentes = result.get('values', [])
        siguiente_fila = len(valores_existentes) + 2
        log(f'Siguiente fila disponible para carga masiva: {siguiente_fila}')
        
        # Escribir desde la siguiente fila vacía
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            values = [[row.get(campo, '')] for row in rows]
            fila_final = siguiente_fila + len(values) - 1
            log(f'Escribiendo campo {campo} en {col}{siguiente_fila}:{col}{fila_final} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{siguiente_fila}:{col}{fila_final}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- BLACKLIST DATOS (OBTENER TODOS) ENDPOINT ---
@app.route('/api/blacklist_datos', methods=['GET'])
def blacklist_datos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        print('[BLACKLIST_DATOS] UID:', uid)
        

        
        # Obtener configuración de hoja/rango
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Blacklist')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[BLACKLIST_DATOS] Ubicaciones:', ubicaciones)
        
        # Obtener URL del spreadsheet
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Blacklist')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        
        # Leer datos de todas las columnas configuradas
        datos = []
        campos_orden = list(ubicaciones.keys())
        
        # Obtener el campo 'Marca' para determinar cuántas filas hay (o el primer campo si no existe)
        primer_campo = 'Marca' if 'Marca' in ubicaciones else campos_orden[0]
        primer_rango = ubicaciones[primer_campo]
        match = re.match(r"([A-Z]+)", primer_rango)
        if match:
            col = match.group(1)
        else:
            col = primer_rango.split(':')[0]
        
        # Usar rango abierto para leer todas las filas disponibles
        rango_lectura = f"{nombre_hoja}!{col}2:{col}"
        print(f'[BLACKLIST_DATOS] Leyendo {rango_lectura} para contar filas (campo: {primer_campo})')
        
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=rango_lectura
        ).execute()
        
        primera_columna = result.get('values', [])
        num_filas = len(primera_columna)
        print(f'[BLACKLIST_DATOS] Filas detectadas: {num_filas}')
        
        if num_filas == 0:
            return jsonify({'datos': [], 'campos': campos_orden})
        
        # Leer todas las columnas usando rangos abiertos
        columnas_datos = {}
        for campo in campos_orden:
            rango = ubicaciones[campo]
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            
            rango_lectura = f"{nombre_hoja}!{col}2:{col}"
            print(f'[BLACKLIST_DATOS] Leyendo {campo} desde {rango_lectura}')
            
            result = sheet.values().get(
                spreadsheetId=spreadsheet_id,
                range=rango_lectura
            ).execute()
            
            valores = result.get('values', [])
            columnas_datos[campo] = [v[0] if v else '' for v in valores]
        
        # Construir array de objetos
        for i in range(num_filas):
            fila = {'_row_index': i + 2}  # +2 porque empezamos en la fila 2
            for campo in campos_orden:
                fila[campo] = columnas_datos[campo][i] if i < len(columnas_datos[campo]) else ''
            datos.append(fila)
        
        print(f'[BLACKLIST_DATOS] Retornando {len(datos)} filas')
        return jsonify({'datos': datos, 'campos': campos_orden})
        
    except Exception as e:
        import traceback
        print('[BLACKLIST_DATOS] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- BLACKLIST ELIMINAR FILA ENDPOINT ---
@app.route('/api/blacklist_eliminar', methods=['POST'])
def blacklist_eliminar():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        row_index = data.get('row_index')
        
        if not row_index:
            return jsonify({'error': 'row_index requerido'}), 400
        
        print(f'[BLACKLIST_ELIMINAR] UID: {uid}, Fila a eliminar: {row_index}')
        

        
        # Obtener configuración de hoja/rango
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('Blacklist')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        
        # Obtener URL del spreadsheet
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('Blacklist')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        # Obtener el sheetId del spreadsheet
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_id = None
        for sheet in spreadsheet.get('sheets', []):
            if sheet['properties']['title'] == nombre_hoja:
                sheet_id = sheet['properties']['sheetId']
                break
        
        if sheet_id is None:
            return jsonify({'error': f'No se encontró la hoja {nombre_hoja}'}), 404
        
        print(f'[BLACKLIST_ELIMINAR] SheetId: {sheet_id}, Eliminando fila {row_index}')
        
        # Eliminar la fila usando batchUpdate
        requests = [{
            'deleteDimension': {
                'range': {
                    'sheetId': sheet_id,
                    'dimension': 'ROWS',
                    'startIndex': row_index - 1,  # 0-indexed
                    'endIndex': row_index  # exclusive
                }
            }
        }]
        
        body = {'requests': requests}
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=body
        ).execute()
        
        print(f'[BLACKLIST_ELIMINAR] Fila {row_index} eliminada exitosamente')
        return jsonify({'status': 'ok', 'deleted_row': row_index})
        
    except Exception as e:
        import traceback
        print('[BLACKLIST_ELIMINAR] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- ENDPOINTS PARA PEDIDOS CALENDARIO ---
@app.route('/api/pedidos_calendario_campos', methods=['GET'])
def pedidos_calendario_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('CALENDARIO')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/api/pedidos_calendario_registro', methods=['POST'])
def pedidos_calendario_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import datetime, re, traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        print('[PEDIDOS_CALENDARIO_REGISTRO] UID:', uid)
        print('[PEDIDOS_CALENDARIO_REGISTRO] Data recibida:', data)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('CALENDARIO')
        print('[PEDIDOS_CALENDARIO_REGISTRO] Spreadsheet URL:', spreadsheet_url)
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        print('[PEDIDOS_CALENDARIO_REGISTRO] Spreadsheet ID:', spreadsheet_id)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('CALENDARIO')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        print('[PEDIDOS_CALENDARIO_REGISTRO] Hoja data:', hoja_data)
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        print('[PEDIDOS_CALENDARIO_REGISTRO] Ubicaciones:', ubicaciones)
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import datetime
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'[PEDIDOS_CALENDARIO_REGISTRO] Limpiando todos los rangos configurados antes de registrar...')
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            print(f'[PEDIDOS_CALENDARIO_REGISTRO] Limpiando rango {clear_range}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=clear_range,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            valor = data.get(campo, '')
            if campo.lower() == 'fecha':
                valor = fecha
            values = [[valor]]
            rango_celda = f"{nombre_hoja}!{col}2:{col}2"
            print(f'[PEDIDOS_CALENDARIO_REGISTRO] Escribiendo campo {campo} en {rango_celda} con valor: {valor}')
            try:
                sheet.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=rango_celda,
                    valueInputOption='USER_ENTERED',
                    body={'values': values}
                ).execute()
            except Exception as e:
                print(f'[PEDIDOS_CALENDARIO_REGISTRO] Error insertando {campo} en {rango_celda}:', str(e))
                print(traceback.format_exc())
                return jsonify({'error': f'Error insertando {campo} en {rango_celda}', 'details': str(e)}), 500
        print('[PEDIDOS_CALENDARIO_REGISTRO] Registro exitoso')
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[PEDIDOS_CALENDARIO_REGISTRO] Error general:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Pedidos Calendario', 'details': str(e)}), 500

@app.route('/api/pedidos_calendario_bulk', methods=['POST'])
def pedidos_calendario_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        import sys
        def log(msg, *args):
            print('[PEDIDOS_CALENDARIO_BULK]', msg, *args)
            sys.stdout.flush()
        log('UID:', uid)

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('CALENDARIO')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            log('No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            log('No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('CALENDARIO')
        if not spreadsheet_url:
            log('No se encontró la URL del spreadsheet para CALENDARIO')
            return jsonify({'error': 'No se encontró la URL del spreadsheet para CALENDARIO. Configura el campo "CALENDARIO" en Firestore.'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            log('URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            log('No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            import csv, io
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError as e:
                log('Fallo utf-8, intentando latin-1:', str(e))
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        import re
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            log(f'Limpiando rango {clear_range}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=clear_range,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            if match:
                col = match.group(1)
            else:
                col = rango.split(':')[0]
            fila = 2
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila + len(values) - 1
            log(f'Escribiendo campo {campo} en {col}{fila}:{col}{fila_final} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila_final}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- ENDPOINT PARA OBTENER DATOS DE CALENDARIO (TABLA A2:G14) ---
@app.route('/api/pedidos_calendario_datos', methods=['GET'])
def pedidos_calendario_datos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener URL del spreadsheet
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404

        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('CALENDARIO')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404

        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)

        # Obtener nombre de hoja
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('CALENDARIO')
        hoja_doc = hoja_ref.get()
        nombre_hoja = 'Sheet1'
        if hoja_doc.exists:
            hoja_data = hoja_doc.to_dict()
            nombre_hoja = hoja_data.get('Hoja', 'Sheet1')

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Leer encabezados de la fila 1
        rango_headers = f"{nombre_hoja}!A1:G1"
        result_headers = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=rango_headers
        ).execute()

        encabezados_raw = result_headers.get('values', [[]])[0] if result_headers.get('values') else []
        encabezados = []
        for i in range(7):
            if i < len(encabezados_raw):
                encabezados.append(encabezados_raw[i])
            else:
                encabezados.append(chr(65 + i))  # A, B, C, etc como fallback

        # Leer rango A2:G14
        rango = f"{nombre_hoja}!A2:G14"
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=rango
        ).execute()

        valores = result.get('values', [])

        # Asegurar que tenemos 13 filas con 7 columnas cada una
        datos = []
        for i in range(13):
            if i < len(valores):
                fila = valores[i]
                # Completar con valores vacíos si faltan columnas
                while len(fila) < 7:
                    fila.append('')
                datos.append(fila[:7])  # Solo tomar primeras 7 columnas
            else:
                datos.append(['', '', '', '', '', '', ''])

        return jsonify({'encabezados': encabezados, 'datos': datos})

    except Exception as e:
        import traceback
        print('[CALENDARIO_DATOS] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- ENDPOINT PARA GUARDAR DATOS DE CALENDARIO (TABLA A2:G14) ---
@app.route('/api/pedidos_calendario_guardar', methods=['POST'])
def pedidos_calendario_guardar():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        tabla_datos = data.get('datos', [])

        # Obtener URL del spreadsheet
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404

        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('CALENDARIO')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404

        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)

        # Obtener nombre de hoja
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('CALENDARIO')
        hoja_doc = hoja_ref.get()
        nombre_hoja = 'Sheet1'
        if hoja_doc.exists:
            hoja_data = hoja_doc.to_dict()
            nombre_hoja = hoja_data.get('Hoja', 'Sheet1')

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # Escribir en rango A2:G14
        rango = f"{nombre_hoja}!A2:G14"
        sheet.values().update(
            spreadsheetId=spreadsheet_id,
            range=rango,
            valueInputOption='USER_ENTERED',
            body={'values': tabla_datos}
        ).execute()

        return jsonify({'status': 'ok'})

    except Exception as e:
        import traceback
        print('[CALENDARIO_GUARDAR] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- ENDPOINTS PARA PEDIDOS BD ---
@app.route('/api/pedidos_bd_campos', methods=['GET'])
def pedidos_bd_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDsi')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/api/pedidos_bd_registro', methods=['POST'])
def pedidos_bd_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import datetime, re, traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDsi')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDsi')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            sheet.values().clear(spreadsheetId=spreadsheet_id, range=clear_range, body={}).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            valor = data.get(campo, '')
            if campo.lower() == 'fecha':
                valor = fecha
            values = [[valor]]
            rango_celda = f"{nombre_hoja}!{col}2:{col}2"
            sheet.values().update(spreadsheetId=spreadsheet_id, range=rango_celda, valueInputOption='USER_ENTERED', body={'values': values}).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[PEDIDOS_BD_REGISTRO] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Pedidos BD', 'details': str(e)}), 500

@app.route('/api/pedidos_bd_bulk', methods=['POST'])
def pedidos_bd_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDsi')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDsi')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet para BDsi'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        rows = []
        if filename.endswith('.csv'):
            import csv, io
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError:
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            if reader.fieldnames is None:
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400

            # Normalizar nombres de columnas (case-insensitive)
            csv_headers_lower = {h.lower(): h for h in reader.fieldnames}
            config_to_csv = {}
            missing = []
            for campo in campos_config:
                campo_lower = campo.lower()
                if campo_lower in csv_headers_lower:
                    config_to_csv[campo] = csv_headers_lower[campo_lower]
                else:
                    missing.append(campo)

            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400

            # Leer filas y mapear nombres de columnas
            rows = []
            for row in reader:
                normalized_row = {}
                for campo_config, campo_csv in config_to_csv.items():
                    normalized_row[campo_config] = row.get(campo_csv, '')
                rows.append(normalized_row)
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Normalizar nombres de columnas (case-insensitive)
            xlsx_headers_lower = {h.lower(): h for h in headers if h}
            config_to_xlsx = {}
            missing = []
            for campo in campos_config:
                campo_lower = campo.lower()
                if campo_lower in xlsx_headers_lower:
                    config_to_xlsx[campo] = xlsx_headers_lower[campo_lower]
                else:
                    missing.append(campo)

            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400

            # Leer filas y mapear nombres de columnas
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                normalized_row = {}
                for campo_config, campo_xlsx in config_to_xlsx.items():
                    normalized_row[campo_config] = row_dict.get(campo_xlsx, '')
                rows.append(normalized_row)
        else:
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            sheet.values().clear(spreadsheetId=spreadsheet_id, range=clear_range, body={}).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            fila = 2
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila + len(values) - 1
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila_final}"
            sheet.values().update(spreadsheetId=spreadsheet_id, range=rango_celda, valueInputOption='USER_ENTERED', body={'values': values}).execute()
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        print('[PEDIDOS_BD_BULK] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ==================== PEDIDOS - BDQTY ====================
@app.route('/api/pedidos_bdqty_campos', methods=['GET'])
def pedidos_bdqty_campos():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Token no proporcionado'}), 401
        id_token = auth_header.split('Bearer ')[1]
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDQTY')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        import traceback
        print('[PEDIDOS_BDQTY_CAMPOS] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/pedidos_bdqty_registro', methods=['POST'])
def pedidos_bdqty_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        import datetime, re, traceback
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDQTY')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet'}), 404
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDQTY')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            sheet.values().clear(spreadsheetId=spreadsheet_id, range=clear_range, body={}).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            valor = data.get(campo, '')
            if campo.lower() == 'fecha':
                valor = fecha
            values = [[valor]]
            rango_celda = f"{nombre_hoja}!{col}2:{col}2"
            sheet.values().update(spreadsheetId=spreadsheet_id, range=rango_celda, valueInputOption='USER_ENTERED', body={'values': values}).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        import traceback
        print('[PEDIDOS_BDQTY_REGISTRO] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': 'Error al registrar en Pedidos BDQTY', 'details': str(e)}), 500

@app.route('/api/pedidos_bdqty_bulk', methods=['POST'])
def pedidos_bdqty_bulk():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Token no proporcionado'}), 401
        id_token = auth_header.split('Bearer ')[1]
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDQTY')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BDQTY')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL del spreadsheet para BDQTY'}), 404
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        rows = []
        if filename.endswith('.csv'):
            import csv, io
            try:
                content = file.stream.read().decode('utf-8')
            except UnicodeDecodeError:
                file.stream.seek(0)
                content = file.stream.read().decode('latin-1')
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            if reader.fieldnames is None:
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400

            # Normalizar nombres de columnas (case-insensitive)
            csv_headers_lower = {h.lower(): h for h in reader.fieldnames}
            config_to_csv = {}
            missing = []
            for campo in campos_config:
                campo_lower = campo.lower()
                if campo_lower in csv_headers_lower:
                    config_to_csv[campo] = csv_headers_lower[campo_lower]
                else:
                    missing.append(campo)

            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400

            # Leer filas y mapear nombres de columnas
            rows = []
            for row in reader:
                normalized_row = {}
                for campo_config, campo_csv in config_to_csv.items():
                    normalized_row[campo_config] = row.get(campo_csv, '')
                rows.append(normalized_row)
        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Normalizar nombres de columnas (case-insensitive)
            xlsx_headers_lower = {h.lower(): h for h in headers if h}
            config_to_xlsx = {}
            missing = []
            for campo in campos_config:
                campo_lower = campo.lower()
                if campo_lower in xlsx_headers_lower:
                    config_to_xlsx[campo] = xlsx_headers_lower[campo_lower]
                else:
                    missing.append(campo)

            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400

            # Leer filas y mapear nombres de columnas
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                normalized_row = {}
                for campo_config, campo_xlsx in config_to_xlsx.items():
                    normalized_row[campo_config] = row_dict.get(campo_xlsx, '')
                rows.append(normalized_row)
        else:
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            clear_range = f"{nombre_hoja}!{col}2:{col}1000"
            sheet.values().clear(spreadsheetId=spreadsheet_id, range=clear_range, body={}).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)", rango)
            col = match.group(1) if match else rango.split(':')[0]
            fila = 2
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila + len(values) - 1
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila_final}"
            sheet.values().update(spreadsheetId=spreadsheet_id, range=rango_celda, valueInputOption='USER_ENTERED', body={'values': values}).execute()
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        print('[PEDIDOS_BDQTY_BULK] Error:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# --- RUTAS PARA COMPRAS ---
@app.route('/compras')
def compras():
    return render_template('compras.html')

@app.route('/actualizar')
def actualizar():
    return render_template('Actualizar.html')

@app.route('/cotizaciones')
def cotizaciones():
    """Muestra la página de visualización de Cotizaciones"""
    return render_template('Cotizaciones.html')

@app.route('/cotizacion_detalle')
def cotizacion_detalle_page():
    """Muestra la página de detalle de una cotización específica"""
    return render_template('CotizacionDetalle.html')

@app.route('/pedidos_anteriores')
def pedidos_anteriores():
    """Muestra la página de visualización de Pedidos Anteriores"""
    return render_template('PedidosAnteriores.html')

@app.route('/indicadores')
def indicadores():
    """Muestra la página de indicadores con los 3 botones"""
    return render_template('indicadores.html')

@app.route('/api/indicadores/<section>')
def get_indicadores_data(section):
    """Obtiene datos filtrados de la hoja de cálculo de indicadores"""
    try:
        from googleapiclient.discovery import build

        # Obtener credenciales
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID de la hoja de cálculo de indicadores
        spreadsheet_id = '1w8OOsQ-N9XkxD84xYIN6XJtjKvgvFBNjIm1U21-ksJk'

        # Obtener ID de la hoja (sheetId) real para operaciones de batchUpdate
        spreadsheet_meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_id = spreadsheet_meta['sheets'][0]['properties']['sheetId']

        # Intentar obtener ID desde Firebase si es necesario, por ahora usamos el hardcoded
        # que coincide con el log de error proporcionado.

        # Leer todos los datos de la hoja
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='A:Z'  # Leer todas las columnas
        ).execute()

        values = result.get('values', [])

        if not values:
            return jsonify({'error': 'No se encontraron datos'}), 404

        # Buscar las palabras clave en la primera columna
        # Cotizados empieza en fila 4 (índice 3) según requerimiento
        cotizados_start = 3 
        surtido_start = -1
        procesado_start = -1
        mesas_start = -1

        for i, row in enumerate(values):
            if row and len(row) > 0:
                cell_value = str(row[0]).upper().strip()
                if 'SURTIDO' in cell_value or cell_value == 'SURTIDO':
                    surtido_start = i
                elif 'PROCESADO' in cell_value or cell_value == 'PROCESADO':
                    procesado_start = i
                elif 'MESAS' in cell_value or cell_value == 'MESAS':
                    mesas_start = i

        # Determinar qué sección devolver
        if section == 'cotizados':
            # Desde fila 4 hasta SURTIDO
            end_index = surtido_start - 1 if surtido_start > cotizados_start else len(values)
            # Incluimos la fila de encabezados si existe, o tomamos desde cotizados_start
            # Asumiendo que los datos empiezan en cotizados_start
            section_data = values[cotizados_start:end_index]

        elif section == 'surtido':
            if surtido_start == -1:
                return jsonify({'error': 'No se encontró la sección SURTIDO'}), 404

            end_index = procesado_start - 1 if procesado_start > surtido_start else len(values)
            section_data = values[surtido_start + 1:end_index]

        elif section == 'procesado':
            if procesado_start == -1:
                return jsonify({'error': 'No se encontró la sección PROCESADO'}), 404

            end_index = mesas_start - 1 if mesas_start > procesado_start else len(values)
            section_data = values[procesado_start + 1:end_index]
        else:
            return jsonify({'error': 'Sección no válida'}), 400

        # Filtrar filas donde la columna B (índice 1) esté vacía
        # Y asegurar que la fila no esté totalmente vacía
        filtered_data = []
        for row in section_data:
            # Verificar si existe columna B y tiene valor
            if len(row) > 1 and row[1].strip() != '':
                filtered_data.append(row)

        # Separar encabezados de datos
        # Asumimos que la primera fila del rango seleccionado NO son encabezados para Cotizados
        # según la descripción "empiezan desde la fila 4". 
        # Si se requieren encabezados fijos, se pueden hardcodear o tomar de la fila 3.
        # Para mantener consistencia con el frontend existente:
        headers = [] 
        rows = filtered_data

        return jsonify({
            'headers': headers,
            'rows': rows,
            'section': section
        })

    except Exception as e:
        print('[INDICADORES] Error:', str(e))
        import traceback
        print(traceback.format_exc())
        if '403' in str(e):
            return jsonify({'error': 'Permiso denegado (403). Por favor comparte la hoja de cálculo con el email de la cuenta de servicio.'}), 403
        return jsonify({'error': str(e)}), 500

@app.route('/api/indicadores/mover_pedido', methods=['POST'])
def mover_pedido_indicadores():
    """Mueve un pedido de una sección a otra en la hoja de indicadores"""
    try:
        from googleapiclient.discovery import build

        data = request.get_json()
        seccion_actual = data.get('seccionActual')
        seccion_destino = data.get('seccionDestino')
        row_data = data.get('rowData')
        row_index = int(data.get('rowIndex')) # Asegurar que sea entero

        if not all([seccion_actual, seccion_destino, row_data]):
            return jsonify({'error': 'Datos incompletos'}), 400

        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        spreadsheet_id = '1w8OOsQ-N9XkxD84xYIN6XJtjKvgvFBNjIm1U21-ksJk'

        # Obtener ID de la hoja (sheetId) real para operaciones de batchUpdate
        spreadsheet_meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_id = spreadsheet_meta['sheets'][0]['properties']['sheetId']

        # Leer todos los datos de la hoja
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='A:Z'
        ).execute()

        values = result.get('values', [])

        # Buscar secciones
        cotizados_start = 3
        surtido_start = -1
        procesado_start = -1
        mesas_start = -1

        for i, row in enumerate(values):
            if row and len(row) > 0:
                cell_value = str(row[0]).upper().strip()
                if 'SURTIDO' in cell_value:
                    surtido_start = i
                elif 'PROCESADO' in cell_value:
                    procesado_start = i
                elif 'MESAS' in cell_value:
                    mesas_start = i

        # Determinar las filas de inicio y fin de cada sección
        sections_map = {
            'cotizados': (cotizados_start, surtido_start - 1 if surtido_start > 0 else len(values)),
            'surtido': (surtido_start + 1 if surtido_start > 0 else -1, procesado_start - 1 if procesado_start > surtido_start else len(values)),
            'procesado': (procesado_start + 1 if procesado_start > 0 else -1, mesas_start - 1 if mesas_start > procesado_start else len(values))
        }

        if seccion_actual not in sections_map:
            return jsonify({'error': 'Sección actual no válida'}), 400

        start_row, end_row = sections_map[seccion_actual]
        if start_row == -1:
            return jsonify({'error': f'No se encontró la sección {seccion_actual}'}), 404

        # Encontrar la fila exacta en la hoja
        # Necesitamos buscar la fila que coincida con row_data
        actual_row_index = -1
        filtered_index = 0
        for i in range(start_row, end_row):
            if i < len(values) and len(values[i]) > 1 and values[i][1].strip():
                if filtered_index == row_index:
                    actual_row_index = i
                    break
                filtered_index += 1

        if actual_row_index == -1:
            return jsonify({'error': 'No se encontró la fila especificada'}), 404

        # CASO 1: Si el destino es "completado", solo eliminar la fila de la sección actual
        if seccion_destino == 'completado':
            # Borrar la fila
            delete_request = {
                'requests': [{
                    'deleteDimension': {
                        'range': {
                            'sheetId': sheet_id,
                            'dimension': 'ROWS',
                            'startIndex': actual_row_index,
                            'endIndex': actual_row_index + 1
                        }
                    }
                }]
            }
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=delete_request
            ).execute()

        else: # CASO 2: Mover a otra sección (Insertar en destino + Borrar de origen)
            # Mover a la nueva sección
            dest_start, _ = sections_map[seccion_destino]
            if dest_start == -1:
                return jsonify({'error': f'No se encontró la sección {seccion_destino}'}), 404

            # Insertar en la nueva sección (después del encabezado)
            insert_row = dest_start + 1

            # Primero insertar una fila vacía en el destino
            insert_request = {
                'requests': [{
                    'insertDimension': {
                        'range': {
                            'sheetId': sheet_id,
                            'dimension': 'ROWS',
                            'startIndex': insert_row,
                            'endIndex': insert_row + 1
                        }
                    }
                }]
            }
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=insert_request
            ).execute()

            # Escribir los datos en la nueva fila
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f'A{insert_row + 1}',  # +1 porque las filas en Sheets empiezan en 1
                valueInputOption='RAW',
                body={'values': [row_data]}
            ).execute()

            # Eliminar la fila original (el índice cambió debido a la inserción)
            # Si insertamos ANTES de la fila original (índice menor), la fila original baja 1 posición.
            # Si insertamos DESPUÉS (índice mayor), la fila original mantiene su índice.
            delete_index = actual_row_index + 1 if insert_row <= actual_row_index else actual_row_index

            delete_request = {
                'requests': [{
                    'deleteDimension': {
                        'range': {
                            'sheetId': sheet_id,
                            'dimension': 'ROWS',
                            'startIndex': delete_index,
                            'endIndex': delete_index + 1
                        }
                    }
                }]
            }
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=delete_request
            ).execute()

        return jsonify({'success': True, 'message': 'Pedido movido exitosamente'})

    except Exception as e:
        print('[MOVER_PEDIDO] Error:', str(e))
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/update_sheet_cell', methods=['POST'])
def update_sheet_cell():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        
        data = request.get_json()
        spreadsheet_id = data.get('spreadsheet_id')
        sheet_name = data.get('sheet_name')
        cell_range = data.get('range') # e.g., 'F7'
        value = data.get('value')

        if not all([spreadsheet_id, sheet_name, cell_range, value is not None]):
            return jsonify({'error': 'Datos incompletos'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        update_range = f"'{sheet_name}'!{cell_range}"

        sheet.values().update(
            spreadsheetId=spreadsheet_id,
            range=update_range,
            valueInputOption='USER_ENTERED',
            body={'values': [[value]]}
        ).execute()

        return jsonify({'status': 'ok', 'message': f'Celda {cell_range} actualizada.'})

    except Exception as e:
        import traceback
        print(f'[UPDATE_SHEET_CELL] Error: {str(e)}')
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/mark_as_unsortable', methods=['POST'])
def mark_as_unsortable():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        
        data = request.get_json()
        spreadsheet_id = data.get('spreadsheet_id')
        sheet_name = data.get('sheet_name')
        row_index = data.get('row_index') # 1-based index

        if not all([spreadsheet_id, sheet_name, row_index]):
            return jsonify({'error': 'Datos incompletos'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()

        # 1. Get sheetId
        spreadsheet_meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_id = None
        for s in spreadsheet_meta.get('sheets', []):
            if s['properties']['title'] == sheet_name:
                sheet_id = s['properties']['sheetId']
                break
        if sheet_id is None:
            return jsonify({'error': f'No se encontró la hoja con nombre {sheet_name}'}), 404

        # 2. Get the row data to be moved
        source_range = f"'{sheet_name}'!A{row_index}:Z{row_index}"
        result = sheet.values().get(spreadsheetId=spreadsheet_id, range=source_range).execute()
        row_values = result.get('values', [[]])[0]

        # 3. Modify data: set "cantidad confirmada" (col F, index 5) to 0
        if len(row_values) > 5:
            row_values[5] = '0'
        else:
            while len(row_values) <= 5:
                row_values.append('')
            row_values[5] = '0'

        # 4. Find the next empty row from row 84
        check_range = f"'{sheet_name}'!A84:A"
        result = sheet.values().get(spreadsheetId=spreadsheet_id, range=check_range).execute()
        existing_values_len = len(result.get('values', []))
        destination_row_index = 84 + existing_values_len

        # 5. Write the modified data to the destination
        destination_range = f"'{sheet_name}'!A{destination_row_index}"
        sheet.values().update(
            spreadsheetId=spreadsheet_id,
            range=destination_range,
            valueInputOption='USER_ENTERED',
            body={'values': [row_values]}
        ).execute()

        # 6. Delete the original row
        delete_request = {
            'requests': [{
                'deleteDimension': {
                    'range': {
                        'sheetId': sheet_id,
                        'dimension': 'ROWS',
                        'startIndex': row_index - 1, # 0-indexed
                        'endIndex': row_index
                    }
                }
            }]
        }
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=delete_request
        ).execute()

        return jsonify({'status': 'ok', 'message': f'Fila {row_index} movida a {destination_row_index}.'})

    except Exception as e:
        import traceback
        print(f'[MARK_UNSORTABLE] Error: {str(e)}')
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/cotizaciones_datos', methods=['GET'])
def cotizaciones_datos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        
        # Obtener URL desde Areas/{uid}
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('COTIZACIONES')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL de COTIZACIONES'}), 404
            
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        # Obtener metadatos para encontrar el nombre correcto de la hoja "Semana"
        spreadsheet_meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet_meta.get('sheets', [])
        sheet_name = 'Semana' # Default
        
        # Buscar hoja que contenga "Semana" (case-insensitive) o usar la primera
        for s in sheets:
            if 'semana' in s['properties']['title'].lower():
                sheet_name = s['properties']['title']
                break
        else:
            if sheets:
                sheet_name = sheets[0]['properties']['title']

        # Leer datos usando el nombre de hoja correcto
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!B2:O22"
        ).execute()

        values = result.get('values', [])

        # Obtener también los datos completos
        result_full = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A1:O100"
        ).execute()
        full_values = result_full.get('values', [])

        # Extraer nombres de días (fila 1: LUNES, MARTES, etc.) y marcas (fila 5 en adelante)
        dias = full_values[0] if len(full_values) > 0 else []
        marcas_por_dia = {}

        # Procesar desde la fila 5 (índice 4) donde empiezan las marcas
        if len(full_values) >= 5:
            for row_idx in range(4, len(full_values)):
                row = full_values[row_idx]
                for col_idx, cell in enumerate(row):
                    if cell and cell.strip():  # Si hay un nombre de marca
                        dia = dias[col_idx] if col_idx < len(dias) else f"Col{col_idx}"
                        if dia not in marcas_por_dia:
                            marcas_por_dia[dia] = []
                        marcas_por_dia[dia].append({
                            'nombre': cell.strip(),
                            'fila': row_idx + 1,
                            'columna': col_idx
                        })

        return jsonify({
            'datos': values,
            'marcas_por_dia': marcas_por_dia,
            'spreadsheet_id': spreadsheet_id,
            'sheet_name': sheet_name
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cotizacion_detalle/<marca>/<semana>', methods=['GET'])
def cotizacion_detalle(marca, semana):
    """Obtiene los detalles de una hoja específica de marca y semana"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener URL desde Areas/{uid}
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404

        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('COTIZACIONES')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL de COTIZACIONES'}), 404

        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Obtener índice de ocurrencia (para duplicados: Marca W3, Marca W3 2, etc.)
        indice = request.args.get('indice', '1')
        
        if int(indice) > 1:
            sheet_name = f"{marca} {semana} {indice}"
        else:
            sheet_name = f"{marca} {semana}"

        # Leer todos los datos de la hoja específica
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A1:Z1000"
        ).execute()

        values = result.get('values', [])
        
        if not values:
             return jsonify({'error': 'HOJA NO EXISTENTE FAVOR DE VERIFICAR MARCA O SEMANA'}), 404

        return jsonify({
            'datos': values,
            'sheet_name': sheet_name,
            'marca': marca,
            'semana': semana
        })

    except Exception as e:
        error_str = str(e)
        # Capturar error específico de rango no encontrado o hoja no existente
        if "Unable to parse range" in error_str or "Not Found" in error_str or "400" in error_str:
             return jsonify({'error': 'HOJA NO EXISTENTE FAVOR DE VERIFICAR MARCA O SEMANA'}), 400
        return jsonify({'error': str(e)}), 500

@app.route('/api/pedidos_anteriores_datos', methods=['GET'])
def pedidos_anteriores_datos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('PEDIDOSANT')
        if not spreadsheet_url:
            return jsonify({'error': 'No se encontró la URL de PEDIDOSANT'}), 404
            
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='A1:Z1000'
        ).execute()
        
        values = result.get('values', [])
        return jsonify({'datos': values})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ejecutar_appscript', methods=['POST'])
def ejecutar_appscript():
    """Ejecuta un AppScript específico según el tipo solicitado usando Web App URLs"""
    try:
        import requests

        data = request.get_json()
        tipo = data.get('tipo')

        if not tipo:
            return jsonify({'error': 'Tipo de script no especificado'}), 400

        # Intentar obtener URLs desde Firebase Realtime Database
        try:
            web_app_urls_ref = firebase_db.reference('appscript_web_urls')
            web_app_urls_firebase = web_app_urls_ref.get() or {}
            print(f'[APPSCRIPT] URLs cargadas desde Firebase Realtime Database: {web_app_urls_firebase}')
        except Exception as e:
            print(f'[APPSCRIPT] Error al leer URLs desde Firebase: {str(e)}')
            web_app_urls_firebase = {}

        # Mapeo de tipos a las URLs de Web App
        # IMPORTANTE: Estas URLs deben ser configuradas en Firebase en la ruta 'appscript_web_urls'
        # O pueden ser actualizadas directamente aquí
        # Para obtener la URL: Implementar > Nueva implementación > Aplicación web > Copiar URL
        scripts_map = {
            'pedidos_anteriores': [
                {
                    'web_app_url': web_app_urls_firebase.get('ExtraergC', 'PENDIENTE_CONFIGURAR_EXTRAERGC'),
                    'name': 'ExtraergC',
                    'function': 'extraerYGuardarDatos'
                },
                {
                    'web_app_url': web_app_urls_firebase.get('ExtraergS', 'PENDIENTE_CONFIGURAR_EXTRAERGS'),
                    'name': 'ExtraergS',
                    'function': 'extraerYGuardarDatos'
                },
                {
                    'web_app_url': web_app_urls_firebase.get('ExtraergSP', 'PENDIENTE_CONFIGURAR_EXTRAERGSP'),
                    'name': 'ExtraergSP',
                    'function': 'extraerYGuardarDatos'
                }
            ],
            'calculadora': [
                {
                    'web_app_url': web_app_urls_firebase.get('generarPedidoFinal', 'PENDIENTE_CONFIGURAR_GENERARPEDIDOFINAL'),
                    'name': 'generarPedidoFinal',
                    'function': 'generarPedidoFinal'
                }
            ],
            'resultados': [
                {
                    'web_app_url': web_app_urls_firebase.get('DupeProXd', 'PENDIENTE_CONFIGURAR_DUPEPROXD'),
                    'name': 'DupeProXd',
                    'function': 'DupeProXd'
                }
            ],
            'creacion_envio': [
                {
                    'web_app_url': web_app_urls_firebase.get('procesoCompleto', 'PENDIENTE_CONFIGURAR_PROCESOCOMPLETO'),
                    'name': 'procesoCompleto',
                    'function': 'procesoCompleto'
                }
            ],
            'indicadores_update': [
                {
                    'web_app_url': web_app_urls_firebase.get('actualizarHoja18DesdeBD', 'PENDIENTE_CONFIGURAR_ACTUALIZARHOJA18'),
                    'name': 'actualizarHoja18DesdeBD',
                    'function': 'actualizarHoja18DesdeBD'
                }
            ],
            'calcular_descuentos': [
                {
                    'web_app_url': web_app_urls_firebase.get('Calculardescuentos', 'PENDIENTE_CONFIGURAR_CALCULARDESCUENTOS'),
                    'name': 'CalcularDescuentos',
                    'function': 'ejecutarProceso'
                }
            ]
        }

        if tipo not in scripts_map:
            return jsonify({'error': 'Tipo de script no válido'}), 400

        scripts = scripts_map[tipo]

        # Ejecutar cada script
        resultados = []
        for script_info in scripts:
            web_app_url = script_info['web_app_url']
            name = script_info['name']
            function_name = script_info.get('function', name)

            # Limpiar comillas extras que puedan venir de Firebase
            if isinstance(web_app_url, str):
                web_app_url = web_app_url.strip().strip('"').strip("'")

            print(f'[APPSCRIPT] Procesando {name} con URL: {web_app_url}')
            print(f'[APPSCRIPT] Función a ejecutar: {function_name}')

            # Verificar si la URL está configurada
            if web_app_url.startswith('PENDIENTE_CONFIGURAR'):
                error_msg = f"URL de Web App no configurada para {name}"
                error_msg += "\n\n💡 PASOS PARA CONFIGURAR:\n"
                error_msg += "1. Abra el script en Google Apps Script\n"
                error_msg += "2. Asegúrese de tener una función doGet() o doPost() en su código\n"
                error_msg += "3. Vaya a: Implementar > Nueva implementación\n"
                error_msg += "4. Seleccione tipo: 'Aplicación web'\n"
                error_msg += "5. Configure:\n"
                error_msg += "   - Descripción: Nombre descriptivo\n"
                error_msg += "   - Ejecutar como: Yo\n"
                error_msg += "   - Quién tiene acceso: Cualquier usuario\n"
                error_msg += "6. Haga clic en 'Implementar'\n"
                error_msg += "7. Copie la URL de la aplicación web\n"
                error_msg += "8. Actualice la configuración en Firebase o en el código Python"

                resultados.append({
                    'script': name,
                    'success': False,
                    'error': error_msg
                })
                print(f'[APPSCRIPT] {error_msg}')
                continue

            print(f'[APPSCRIPT] Ejecutando {name} via Web App...')

            # Para scripts de descuentos, ejecutar en segundo plano y retornar inmediatamente
            if tipo == 'calcular_descuentos':
                import threading

                def ejecutar_en_segundo_plano():
                    try:
                        params = {'func': function_name}
                        print(f'[APPSCRIPT ASYNC] Iniciando ejecución de {name}...')
                        response = requests.get(web_app_url, params=params, timeout=1800)
                        print(f'[APPSCRIPT ASYNC] {name} completado con código: {response.status_code}')
                        if response.status_code == 200:
                            print(f'[APPSCRIPT ASYNC] Respuesta: {response.text[:200]}')
                    except Exception as e:
                        print(f'[APPSCRIPT ASYNC] Error en {name}: {str(e)}')

                # Iniciar thread en segundo plano
                thread = threading.Thread(target=ejecutar_en_segundo_plano)
                thread.daemon = True
                thread.start()

                # Retornar inmediatamente al frontend
                resultados.append({
                    'script': name,
                    'success': True,
                    'result': 'Proceso iniciado en segundo plano. Verifica la hoja de cálculo en 1-2 minutos.',
                    'async': True
                })
                print(f'[APPSCRIPT] {name} iniciado en segundo plano')
                continue

            try:
                # Hacer GET request a la Web App con el nombre de la función como parámetro
                # Timeout de 1800 segundos (30 minutos) para scripts largos
                params = {'func': function_name}
                print(f'[APPSCRIPT] Esperando respuesta (timeout: 1800s)...')
                response = requests.get(web_app_url, params=params, timeout=1800)
                print(f'[APPSCRIPT] Respuesta recibida con código: {response.status_code}')

                if response.status_code == 200:
                    # Intentar parsear como JSON
                    try:
                        result_data = response.json()
                        print(f'[APPSCRIPT] Respuesta de {name}: {result_data}')

                        if isinstance(result_data, dict) and 'error' in result_data:
                            resultados.append({
                                'script': name,
                                'success': False,
                                'error': result_data.get('error', 'Error desconocido'),
                                'details': result_data
                            })
                            print(f'[APPSCRIPT] Error en {name}: {result_data.get("error")}')
                        else:
                            resultados.append({
                                'script': name,
                                'success': True,
                                'result': result_data
                            })
                            print(f'[APPSCRIPT] {name} ejecutado exitosamente')
                            if isinstance(result_data, dict) and 'data' in result_data:
                                print(f'[APPSCRIPT] Detalles: {result_data["data"]}')
                    except Exception as parse_error:
                        # Si no es JSON, puede ser un error HTML
                        print(f'[APPSCRIPT] No se pudo parsear JSON: {parse_error}')

                        # Si es HTML de error, extraer el mensaje
                        if '<html>' in response.text.lower() or '<!doctype html>' in response.text.lower():
                            print(f'[APPSCRIPT] ERROR: Respuesta HTML recibida (Apps Script devolvió error)')
                            print(f'[APPSCRIPT] HTML completo: {response.text[:1500]}')

                            resultados.append({
                                'script': name,
                                'success': False,
                                'error': 'Apps Script devolvió una página HTML de error en lugar de JSON. Revisa los permisos del script y verifica que la función se ejecute correctamente.',
                                'html_snippet': response.text[:500]
                            })
                        else:
                            print(f'[APPSCRIPT] Respuesta texto: {response.text[:500]}')
                            resultados.append({
                                'script': name,
                                'success': True,
                                'result': response.text
                            })
                            print(f'[APPSCRIPT] {name} ejecutado exitosamente')
                else:
                    error_msg = f'HTTP {response.status_code}: {response.text}'
                    resultados.append({
                        'script': name,
                        'success': False,
                        'error': error_msg
                    })
                    print(f'[APPSCRIPT] Error en {name}: {error_msg}')

            except requests.exceptions.Timeout:
                error_msg = 'La ejecución del script superó el tiempo límite (10 minutos)'
                resultados.append({
                    'script': name,
                    'success': False,
                    'error': error_msg
                })
                print(f'[APPSCRIPT] Error en {name}: {error_msg}')
            except Exception as script_error:
                error_msg = f'Error al ejecutar: {str(script_error)}'
                resultados.append({
                    'script': name,
                    'success': False,
                    'error': error_msg
                })
                print(f'[APPSCRIPT] Error en {name}: {error_msg}')

        # Verificar si todos fueron exitosos
        todos_exitosos = all(r['success'] for r in resultados)

        if todos_exitosos:
            return jsonify({
                'success': True,
                'message': f'Todos los scripts de {tipo} ejecutados exitosamente',
                'resultados': resultados
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Algunos scripts de {tipo} fallaron',
                'resultados': resultados
            }), 500

    except Exception as e:
        print('[EJECUTAR_APPSCRIPT] Error:', str(e))
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/proxy_spreadsheet_data', methods=['POST'])
def proxy_spreadsheet_data():
    """Endpoint para obtener datos de una hoja externa dado su URL (para visualización en HTML)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        url = data.get('url')
        if not url:
             return jsonify({'error': 'URL required'}), 400
             
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
        if not match:
            return jsonify({'error': 'Invalid URL'}), 400
        spreadsheet_id = match.group(1)
        
        # Intentar extraer el gid de la URL
        gid_match = re.search(r"[#&?]gid=([0-9]+)", url)
        target_gid = int(gid_match.group(1)) if gid_match else None
        
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        # Obtener metadatos del spreadsheet
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        if not sheets:
             return jsonify({'error': 'No sheets found'}), 404
        
        sheet_name = None
        # Buscar la hoja por gid si existe
        if target_gid is not None:
            for s in sheets:
                if s['properties']['sheetId'] == target_gid:
                    sheet_name = s['properties']['title']
                    break
        
        # Si no se encontró por gid o no había gid, usar la primera
        if not sheet_name:
            sheet_name = sheets[0]['properties']['title']
        
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A1:Z1000"
        ).execute()
        
        return jsonify({'values': result.get('values', []), 'sheetName': sheet_name, 'spreadsheet_id': spreadsheet_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- Nuevas rutas para Descuentos ---
@app.route('/descuentos')
def descuentos():
    return render_template('descuentos.html')

@app.route('/bd_descuentos')
def bd_descuentos():
    return render_template('bd_descuentos.html')

@app.route('/metricas_productos')
def metricas_productos():
    return render_template('metricas_productos.html')

# --- API Endpoint para METRICAS PRODUCTOS ---
@app.route('/api/metricas_productos/data', methods=['POST'])
def metricas_productos_data():
    """Obtiene datos de las hojas de WOWITEMS para METRICAS PRODUCTOS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)

        data = request.get_json()
        sheet_name = data.get('sheet', '')
        range_str = data.get('range', '')

        if not sheet_name or not range_str:
            return jsonify({'error': 'Faltan parámetros sheet y range'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # WOWITEMS spreadsheet ID
        spreadsheet_id = '18rvbsRrqPcZP8W5AYJyc_ivxPIRFl46b2tAyq6Fvv9g'

        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!{range_str}"
        ).execute()

        return jsonify({'values': result.get('values', [])}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# --- ENDPOINTS PARA HISTORICO ---

# Configuracion por defecto de HISTORICO
HISTORICO_DEFAULT_URL = 'https://docs.google.com/spreadsheets/d/13tQ5dqMGuT7M9cYOQdvQozNu7fxW4RUc2-1LNR_nHB0/edit?gid=549893191#gid=549893191'
HISTORICO_DEFAULT_HOJA = {
    'Hoja': 'BD PUBLICACIONES',
    'ID': 'A:A',
    'VENTAS TOTALES': 'H:H'
}

def _init_historico_config(uid):
    """Auto-crea la configuracion de HISTORICO en Firebase si no existe"""
    area_ref = db.collection('Areas').document(uid)
    area_doc = area_ref.get()
    if not area_doc.exists:
        area_ref.set({'HISTORICO': HISTORICO_DEFAULT_URL})
    else:
        area_data = area_doc.to_dict()
        if not area_data.get('HISTORICO'):
            area_ref.update({'HISTORICO': HISTORICO_DEFAULT_URL})
    hoja_ref = area_ref.collection('Hojas').document('BDPUBLICACIONES')
    hoja_doc = hoja_ref.get()
    if not hoja_doc.exists:
        hoja_ref.set(HISTORICO_DEFAULT_HOJA)

@app.route('/api/historico_campos', methods=['GET'])
def historico_campos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        _init_historico_config(uid)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDPUBLICACIONES')
        hoja_doc = hoja_ref.get()
        hoja_data = hoja_doc.to_dict()
        campos = [k for k in hoja_data.keys() if k != 'Hoja']
        return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': 'Error al obtener campos', 'details': str(e)}), 500

@app.route('/api/historico_registro', methods=['POST'])
def historico_registro():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import re
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        _init_historico_config(uid)
        data = request.get_json()
        area_doc = db.collection('Areas').document(uid).get()
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('HISTORICO')
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet invalida'}), 400
        spreadsheet_id = match.group(1)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDPUBLICACIONES')
        hoja_doc = hoja_ref.get()
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'BD PUBLICACIONES')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        # Obtener la ultima fila usada para agregar datos al final
        first_campo = list(ubicaciones.keys())[0]
        first_rango = ubicaciones[first_campo]
        col_letter = first_rango.split(':')[0]
        col_only = ''.join(filter(str.isalpha, col_letter))
        result_range = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=f"{nombre_hoja}!{col_only}:{col_only}"
        ).execute()
        existing_rows = result_range.get('values', [])
        next_row = len(existing_rows) + 1
        # Escribir cada campo en la siguiente fila disponible
        for campo, rango in ubicaciones.items():
            col = ''.join(filter(str.isalpha, rango.split(':')[0]))
            valor = data.get(campo, '')
            rango_celda = f"{nombre_hoja}!{col}{next_row}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': [[valor]]}
            ).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/historico_bulk', methods=['POST'])
def historico_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import re, io, csv, openpyxl
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        _init_historico_config(uid)
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BDPUBLICACIONES')
        hoja_doc = hoja_ref.get()
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'BD PUBLICACIONES')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        area_doc = db.collection('Areas').document(uid).get()
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('HISTORICO')
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet invalida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        rows = []
        if filename.endswith('.csv'):
            try:
                stream = io.StringIO(file.stream.read().decode('utf-8'))
            except Exception:
                stream = io.StringIO(file.stream.read().decode('latin-1'))
            reader = csv.DictReader(stream)
            if reader.fieldnames is None:
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=rango_col,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            match = re.match(r"([A-Z]+)(\d+):", rango)
            if match:
                col = match.group(1)
                fila_inicial = int(match.group(2))
            else:
                col = rango.split(':')[0]
                fila_inicial = 1
            values = [[row.get(campo, '')] for row in rows]
            fila_final = fila_inicial + len(values) - 1
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=f"{nombre_hoja}!{col}{fila_inicial}:{col}{fila_final}",
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- Sub-rutas para BD Descuentos ---
@app.route('/<page_name>')
def bd_descuentos_page(page_name):
    if page_name in ['ventas_semanales', 'para_impulsar', 'para_descartar', 'para_poner_en_venta', 'promocionables', 'errores', 'porcentajes', 'corregir']:
        return render_template(f'{page_name}.html')
    return "Página no encontrada", 404

# --- API Endpoints para Sistema de Descuentos ---

# VENTAS SEMANALES Endpoints
@app.route('/api/ventas_semanales/add', methods=['POST'])
def ventas_semanales_add():
    """Añade un registro individual a VENTAS SEMANALES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        sku = data.get('sku', '').strip()
        unidades = data.get('unidades', '').strip()
        precio_venta = data.get('precio_venta', '').strip()

        if not sku or not unidades or not precio_venta:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # === PRIMERA HOJA: BDPROMOTE ===
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'
        sheet_name = 'VENTAS SEMANALES'

        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = len(existing_rows) + 1

        # SKU va a A y D, Unidades va a I
        batch_data = [
            {'range': f'{sheet_name}!A{next_row}', 'values': [[sku]]},
            {'range': f'{sheet_name}!D{next_row}', 'values': [[sku]]},
            {'range': f'{sheet_name}!I{next_row}', 'values': [[unidades]]}
        ]

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        # === SEGUNDA HOJA: BDVENTASWOWITEMS ===
        spreadsheet_id_2 = '1yMGIpgtfnz1ROcJ0d9dzGBsdxPLwBCt_danE8S8AA6U'
        sheet_name_2 = 'BD VENTAS FULL WEEK'

        result_range_2 = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id_2,
            range=f"'{sheet_name_2}'!Q:Q"
        ).execute()

        existing_rows_2 = result_range_2.get('values', [])
        next_row_2 = max(len(existing_rows_2) + 1, 2)  # Mínimo fila 2

        # SKU va a Q y R, Unidades va a G, Precio de venta va a H
        batch_data_2 = [
            {'range': f"'{sheet_name_2}'!Q{next_row_2}", 'values': [[sku]]},
            {'range': f"'{sheet_name_2}'!R{next_row_2}", 'values': [[sku]]},
            {'range': f"'{sheet_name_2}'!G{next_row_2}", 'values': [[unidades]]},
            {'range': f"'{sheet_name_2}'!H{next_row_2}", 'values': [[precio_venta]]}
        ]

        batch_body_2 = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data_2
        }

        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id_2,
            body=batch_body_2
        ).execute()

        return jsonify({'message': 'Registro añadido exitosamente en ambas hojas'}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ventas_semanales/bulk', methods=['POST'])
def ventas_semanales_bulk():
    """Añade múltiples registros a VENTAS SEMANALES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'error': 'No hay datos para procesar'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Validar datos
        valid_rows = []
        for row in rows:
            sku = str(row.get('sku', '')).strip()
            unidades = str(row.get('unidades', '')).strip()
            precio_venta = str(row.get('precio_venta', '')).strip()
            if sku and unidades:
                valid_rows.append({'sku': sku, 'unidades': unidades, 'precio_venta': precio_venta})

        if not valid_rows:
            return jsonify({'error': 'No hay datos válidos para procesar'}), 400

        # === PRIMERA HOJA: BDPROMOTE ===
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'
        sheet_name = 'VENTAS SEMANALES'

        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = len(existing_rows) + 1

        # Preparar batch update para BDPROMOTE
        batch_data = []
        for i, row_data in enumerate(valid_rows):
            current_row = next_row + i
            batch_data.append({'range': f'{sheet_name}!A{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!D{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!I{current_row}', 'values': [[row_data['unidades']]]})

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        # === SEGUNDA HOJA: BDVENTASWOWITEMS ===
        spreadsheet_id_2 = '1yMGIpgtfnz1ROcJ0d9dzGBsdxPLwBCt_danE8S8AA6U'
        sheet_name_2 = 'BD VENTAS FULL WEEK'

        result_range_2 = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id_2,
            range=f"'{sheet_name_2}'!Q:Q"
        ).execute()

        existing_rows_2 = result_range_2.get('values', [])
        next_row_2 = max(len(existing_rows_2) + 1, 2)  # Mínimo fila 2

        # Preparar batch update para BDVENTASWOWITEMS
        batch_data_2 = []
        for i, row_data in enumerate(valid_rows):
            current_row_2 = next_row_2 + i
            batch_data_2.append({'range': f"'{sheet_name_2}'!Q{current_row_2}", 'values': [[row_data['sku']]]})
            batch_data_2.append({'range': f"'{sheet_name_2}'!R{current_row_2}", 'values': [[row_data['sku']]]})
            batch_data_2.append({'range': f"'{sheet_name_2}'!G{current_row_2}", 'values': [[row_data['unidades']]]})
            if row_data['precio_venta']:
                batch_data_2.append({'range': f"'{sheet_name_2}'!H{current_row_2}", 'values': [[row_data['precio_venta']]]})

        batch_body_2 = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data_2
        }

        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id_2,
            body=batch_body_2
        ).execute()

        return jsonify({'message': f'{len(valid_rows)} registros añadidos exitosamente en ambas hojas'}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# PARA IMPULSAR Endpoints
@app.route('/api/para_impulsar/add', methods=['POST'])
def para_impulsar_add():
    """Añade un registro individual a PARA IMPULSAR VENTAS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        sku = data.get('sku', '').strip()
        unidades_impulsar = data.get('unidades_impulsar', '').strip()
        ventas_30dias = data.get('ventas_30dias', '').strip()

        if not sku or not unidades_impulsar or not ventas_30dias:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Obtener la última fila - empezar desde fila 2
        sheet_name = 'PARA IMPULSAR VENTAS'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 2)  # Mínimo fila 2

        # SKU va a C y D, Ventas 30 días va a I, Unidades impulsar va a L
        batch_data = [
            {
                'range': f'{sheet_name}!C{next_row}',
                'values': [[sku]]
            },
            {
                'range': f'{sheet_name}!D{next_row}',
                'values': [[sku]]
            },
            {
                'range': f'{sheet_name}!I{next_row}',
                'values': [[ventas_30dias]]
            },
            {
                'range': f'{sheet_name}!L{next_row}',
                'values': [[unidades_impulsar]]
            }
        ]

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': 'Registro añadido exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/para_impulsar/bulk', methods=['POST'])
def para_impulsar_bulk():
    """Añade múltiples registros a PARA IMPULSAR VENTAS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'error': 'No hay datos para procesar'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Validar datos
        valid_rows = []
        for row in rows:
            sku = row.get('sku', '').strip()
            unidades_impulsar = row.get('unidades_impulsar', '').strip()
            ventas_30dias = row.get('ventas_30dias', '').strip()
            if sku and unidades_impulsar and ventas_30dias:
                valid_rows.append({'sku': sku, 'unidades_impulsar': unidades_impulsar, 'ventas_30dias': ventas_30dias})

        if not valid_rows:
            return jsonify({'error': 'No hay datos válidos para procesar'}), 400

        # Obtener la última fila - empezar desde fila 2
        sheet_name = 'PARA IMPULSAR VENTAS'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 2)  # Mínimo fila 2

        # Preparar batch update
        batch_data = []
        for i, row_data in enumerate(valid_rows):
            current_row = next_row + i
            # SKU a C y D, Ventas 30 días a I, Unidades impulsar a L
            batch_data.append({'range': f'{sheet_name}!C{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!D{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!I{current_row}', 'values': [[row_data['ventas_30dias']]]})
            batch_data.append({'range': f'{sheet_name}!L{current_row}', 'values': [[row_data['unidades_impulsar']]]})

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': f'{len(valid_rows)} registros añadidos exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# PARA DESCARTAR Endpoints
@app.route('/api/para_descartar/add', methods=['POST'])
def para_descartar_add():
    """Añade un registro individual a PARA EVITAR DESCARTE"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        sku = data.get('sku', '').strip()

        if not sku:
            return jsonify({'error': 'Falta el SKU'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Get last row - empezar desde fila 3
        sheet_name = 'PARA EVITAR DESCARTE'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 3)  # Mínimo fila 3

        # Use batchUpdate for specific columns: SKU en C y D
        batch_data = [
            {'range': f'{sheet_name}!C{next_row}', 'values': [[sku]]},
            {'range': f'{sheet_name}!D{next_row}', 'values': [[sku]]}
        ]

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': 'Registro añadido exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/para_descartar/bulk', methods=['POST'])
def para_descartar_bulk():
    """Añade múltiples registros a PARA EVITAR DESCARTE"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'error': 'No hay datos para procesar'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Validar datos
        valid_rows = []
        for row in rows:
            sku = row.get('sku', '').strip()
            if sku:
                valid_rows.append({'sku': sku})

        if not valid_rows:
            return jsonify({'error': 'No hay datos válidos para procesar'}), 400

        # Get last row - empezar desde fila 3
        sheet_name = 'PARA EVITAR DESCARTE'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 3)  # Mínimo fila 3

        # Prepare batch update for all rows
        batch_data = []
        for i, row_data in enumerate(valid_rows):
            current_row = next_row + i
            batch_data.append({'range': f'{sheet_name}!C{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!D{current_row}', 'values': [[row_data['sku']]]})

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': f'{len(valid_rows)} registros añadidos exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# PARA PONER EN VENTA Endpoints
@app.route('/api/para_poner_en_venta/add', methods=['POST'])
def para_poner_en_venta_add():
    """Añade un registro individual a PARA PONER EN VENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        sku = data.get('sku', '').strip()

        if not sku:
            return jsonify({'error': 'Falta el SKU'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Get last row - empezar desde fila 2
        sheet_name = 'PARA PONER EN VENTA'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 2)  # Mínimo fila 2

        # Use batchUpdate for specific columns: SKU en C y D
        batch_data = [
            {'range': f'{sheet_name}!C{next_row}', 'values': [[sku]]},
            {'range': f'{sheet_name}!D{next_row}', 'values': [[sku]]}
        ]

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': 'Registro añadido exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/para_poner_en_venta/bulk', methods=['POST'])
def para_poner_en_venta_bulk():
    """Añade múltiples registros a PARA PONER EN VENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'error': 'No hay datos para procesar'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet BDPROMOTE
        spreadsheet_id = '14F6ZSyrhp9_f6tHYz6GYaIVqoAEdZo6UICJP0_GR7ew'

        # Validar datos
        valid_rows = []
        for row in rows:
            sku = row.get('sku', '').strip()
            if sku:
                valid_rows.append({'sku': sku})

        if not valid_rows:
            return jsonify({'error': 'No hay datos válidos para procesar'}), 400

        # Get last row - empezar desde fila 2
        sheet_name = 'PARA PONER EN VENTA'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = max(len(existing_rows) + 1, 2)  # Mínimo fila 2

        # Prepare batch update for all rows
        batch_data = []
        for i, row_data in enumerate(valid_rows):
            current_row = next_row + i
            batch_data.append({'range': f'{sheet_name}!C{current_row}', 'values': [[row_data['sku']]]})
            batch_data.append({'range': f'{sheet_name}!D{current_row}', 'values': [[row_data['sku']]]})

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': f'{len(valid_rows)} registros añadidos exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# PROMOCIONABLES Endpoints (solo lectura)
@app.route('/api/promocionables/data', methods=['GET'])
def promocionables_data():
    """Obtiene los datos de la tabla PROMOCIONABLES (PROMOTE 5.0)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet PROMOTE5.0
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Leer los datos de PROMOTE 5.0
        # Columnas: G (% DESCUENTO), A (CATEGORIA), B (ID), C (MARCA), H (PRECIO),
        #           I (PRECIO OFERTA), L (RANGO), K (UTILIDAD), J (UTILIDAD MONEDA)
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='PROMOTE 5.0!A2:L'
        ).execute()

        values = result.get('values', [])

        # Formatear los datos
        data = []
        for row in values:
            # Asegurar que la fila tenga suficientes columnas
            while len(row) < 12:
                row.append('')

            data.append({
                'categoria': row[0] if len(row) > 0 else '',
                'id': row[1] if len(row) > 1 else '',
                'marca': row[2] if len(row) > 2 else '',
                'descuento': row[6] if len(row) > 6 else '',
                'precio': row[7] if len(row) > 7 else '',
                'precio_oferta': row[8] if len(row) > 8 else '',
                'utilidad_moneda': row[9] if len(row) > 9 else '',
                'utilidad': row[10] if len(row) > 10 else '',
                'rango': row[11] if len(row) > 11 else ''
            })

        return jsonify({'data': data}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ERRORES Endpoints (solo lectura + plantillas)
@app.route('/api/errores/data', methods=['GET'])
def errores_data():
    """Obtiene los datos de la tabla ERRORES (ID ERROR)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet PROMOTE5.0
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Leer los datos de ID ERROR
        # Columnas: A (ID), B (COSTO), C (COMISION), D (ENVIO)
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='ID ERROR!A2:D'
        ).execute()

        values = result.get('values', [])

        # Formatear los datos
        data = []
        for row in values:
            # Asegurar que la fila tenga suficientes columnas
            while len(row) < 4:
                row.append('')

            data.append({
                'id': row[0] if len(row) > 0 else '',
                'costo': row[1] if len(row) > 1 else '',
                'comision': row[2] if len(row) > 2 else '',
                'envio': row[3] if len(row) > 3 else ''
            })

        return jsonify({'data': data}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/errores/plantilla_csv', methods=['GET'])
def errores_plantilla_csv():
    """Descarga plantilla CSV con errores marcados"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet PROMOTE5.0
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Leer los datos de ID ERROR
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='ID ERROR!A2:D'
        ).execute()

        values = result.get('values', [])

        # Crear CSV con marcas de error
        # Plantilla debe tener: SKU, MARCA, COSTO, PRECIO, ¿ENVIO?, ENVIO
        from flask import Response
        import io

        output = io.StringIO()
        output.write('SKU,MARCA,COSTO,PRECIO,¿ENVIO?,ENVIO\n')

        for row in values:
            # Asegurar que la fila tenga 4 columnas
            while len(row) < 4:
                row.append('')

            # Solo marcar ERROR donde realmente dice ERROR en la hoja
            # SKU = ID (columna A)
            sku_val = '[ERROR]' if row[0] == 'ERROR' or not row[0] else row[0]
            marca_val = ''  # No está en la hoja ID ERROR
            # COSTO = COSTO (columna B)
            costo_val = '[ERROR]' if row[1] == 'ERROR' or not row[1] else row[1]
            precio_val = ''  # No está en la hoja ID ERROR
            tiene_envio_val = ''  # No está en la hoja ID ERROR
            # ENVIO = ENVIO (columna D)
            envio_val = '[ERROR]' if row[3] == 'ERROR' or not row[3] else row[3]

            output.write(f'{sku_val},{marca_val},{costo_val},{precio_val},{tiene_envio_val},{envio_val}\n')

        csv_content = output.getvalue()
        output.close()

        return Response(
            csv_content,
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=plantilla_errores.csv'}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/errores/plantilla_excel', methods=['GET'])
def errores_plantilla_excel():
    """Descarga plantilla Excel con celdas rojas para errores"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet PROMOTE5.0
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Leer los datos de ID ERROR
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='ID ERROR!A2:D'
        ).execute()

        values = result.get('values', [])

        # Crear Excel con formato
        # Plantilla debe tener: SKU, MARCA, COSTO, PRECIO, ¿ENVIO?, ENVIO
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from flask import Response
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = 'Errores'

        # Encabezados
        ws.append(['SKU', 'MARCA', 'COSTO', 'PRECIO', '¿ENVIO?', 'ENVIO'])

        # Estilo para errores (fondo rojo)
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        white_font = Font(color='FFFFFFFF', bold=True)

        # Agregar datos
        for row in values:
            # Asegurar que la fila tenga 4 columnas
            while len(row) < 4:
                row.append('')

            # Solo marcar ERROR donde realmente dice ERROR en la hoja
            # SKU = ID (columna A)
            sku_val = row[0] if row[0] else ''
            sku_is_error = row[0] == 'ERROR' or not row[0]

            marca_val = ''

            # COSTO = COSTO (columna B)
            costo_val = row[1] if row[1] else ''
            costo_is_error = row[1] == 'ERROR' or not row[1]

            precio_val = ''
            tiene_envio_val = ''

            # ENVIO = ENVIO (columna D)
            envio_val = row[3] if row[3] else ''
            envio_is_error = row[3] == 'ERROR' or not row[3]

            # Agregar fila con valores originales o vacíos
            ws.append([
                'ERROR' if sku_is_error else sku_val,
                marca_val,
                'ERROR' if costo_is_error else costo_val,
                precio_val,
                tiene_envio_val,
                'ERROR' if envio_is_error else envio_val
            ])
            current_row = ws.max_row

            # Marcar solo las celdas que tienen ERROR en rojo
            if sku_is_error:
                cell = ws.cell(row=current_row, column=1)
                cell.fill = red_fill
                cell.font = white_font
                cell.value = 'ERROR'

            if costo_is_error:
                cell = ws.cell(row=current_row, column=3)
                cell.fill = red_fill
                cell.font = white_font
                cell.value = 'ERROR'

            if envio_is_error:
                cell = ws.cell(row=current_row, column=6)
                cell.fill = red_fill
                cell.font = white_font
                cell.value = 'ERROR'

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=plantilla_errores.xlsx'}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# CORREGIR Endpoints
@app.route('/api/corregir/add', methods=['POST'])
def corregir_add():
    """Añade un registro individual a BD PROMOTE"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        sku = data.get('sku', '').strip()
        marca = data.get('marca', '').strip()
        costo = data.get('costo', '').strip()
        precio = data.get('precio', '').strip()
        tiene_envio = data.get('tiene_envio', '').strip()
        envio = data.get('envio', '').strip()

        if not sku or not marca or not costo or not precio or not tiene_envio or not envio:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        spreadsheet_id = '1nPFaXwMqBuFKJpSps85rYTu5gQhwB2dAzf_l2bXtEis'

        # Get last row
        sheet_name = 'BD PROMOTE'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = len(existing_rows) + 1

        # Use batchUpdate for specific columns
        # A (ID=SKU), B (SKU), C (MARCA), D (COSTO), E (PRECIO), G (¿ENVIO?), H (ENVIO)
        batch_data = [
            {'range': f'{sheet_name}!A{next_row}', 'values': [[sku]]},  # ID = SKU
            {'range': f'{sheet_name}!B{next_row}', 'values': [[sku]]},  # SKU
            {'range': f'{sheet_name}!C{next_row}', 'values': [[marca]]},
            {'range': f'{sheet_name}!D{next_row}', 'values': [[costo]]},
            {'range': f'{sheet_name}!E{next_row}', 'values': [[precio]]},
            {'range': f'{sheet_name}!G{next_row}', 'values': [[tiene_envio]]},
            {'range': f'{sheet_name}!H{next_row}', 'values': [[envio]]}
        ]

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': 'Registro añadido exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/corregir/bulk', methods=['POST'])
def corregir_bulk():
    """Añade múltiples registros a BD PROMOTE"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'error': 'No hay datos para procesar'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        spreadsheet_id = '1nPFaXwMqBuFKJpSps85rYTu5gQhwB2dAzf_l2bXtEis'

        # Validar datos
        valid_rows = []
        for row in rows:
            sku = row.get('sku', '').strip()
            marca = row.get('marca', '').strip()
            costo = row.get('costo', '').strip()
            precio = row.get('precio', '').strip()
            tiene_envio = row.get('tiene_envio', '').strip()
            envio = row.get('envio', '').strip()

            if sku and marca and costo and precio and tiene_envio and envio:
                valid_rows.append({
                    'sku': sku,
                    'marca': marca,
                    'costo': costo,
                    'precio': precio,
                    'tiene_envio': tiene_envio,
                    'envio': envio
                })

        if not valid_rows:
            return jsonify({'error': 'No hay filas válidas para procesar'}), 400

        # Get last row
        sheet_name = 'BD PROMOTE'
        result_range = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A:A'
        ).execute()

        existing_rows = result_range.get('values', [])
        next_row = len(existing_rows) + 1

        # Prepare batch update for all rows
        batch_data = []
        for i, row_data in enumerate(valid_rows):
            current_row = next_row + i
            batch_data.append({'range': f'{sheet_name}!A{current_row}', 'values': [[row_data['sku']]]})  # ID
            batch_data.append({'range': f'{sheet_name}!B{current_row}', 'values': [[row_data['sku']]]})  # SKU
            batch_data.append({'range': f'{sheet_name}!C{current_row}', 'values': [[row_data['marca']]]})
            batch_data.append({'range': f'{sheet_name}!D{current_row}', 'values': [[row_data['costo']]]})
            batch_data.append({'range': f'{sheet_name}!E{current_row}', 'values': [[row_data['precio']]]})
            batch_data.append({'range': f'{sheet_name}!G{current_row}', 'values': [[row_data['tiene_envio']]]})
            batch_data.append({'range': f'{sheet_name}!H{current_row}', 'values': [[row_data['envio']]]})

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': f'{len(valid_rows)} registros añadidos exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# PORCENTAJES Endpoints
@app.route('/api/porcentajes/data', methods=['GET'])
def porcentajes_data():
    """Obtiene los datos de la tabla PORCENTAGE TABLE"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Leer los datos desde A2 (CATEGORIA), B2 (DESCUENTO MAXIMO), C2 (UTILIDAD MINIMA)
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='PORCENTAGE TABLE!A2:C'
        ).execute()

        values = result.get('values', [])

        # Formatear los datos
        data = []
        for row in values:
            # Asegurar que la fila tenga suficientes columnas
            while len(row) < 3:
                row.append('')

            data.append({
                'categoria': row[0] if len(row) > 0 else '',
                'descuento_maximo': row[1] if len(row) > 1 else '',
                'utilidad_minima': row[2] if len(row) > 2 else ''
            })

        return jsonify({'data': data}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/porcentajes/update', methods=['POST'])
def porcentajes_update():
    """Actualiza los datos de la tabla PORCENTAGE TABLE (solo columnas B y C)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401

    id_token = auth_header.split(' ')[1]

    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']

        data = request.get_json()
        rows = data.get('data', [])

        if not rows:
            return jsonify({'error': 'No hay datos para actualizar'}), 400

        # Obtener credenciales y conectar con Google Sheets
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # ID del spreadsheet
        spreadsheet_id = '1nOB3Lr07FqOcFKj-1dxNCRnUxGOjUoqAf_WterqH3rg'

        # Preparar batch update solo para columnas B y C
        batch_data = []
        for i, row_data in enumerate(rows):
            row_num = i + 2  # Empezar desde fila 2
            # Solo actualizar columna B (DESCUENTO MAXIMO) y C (UTILIDAD MINIMA)
            batch_data.append({
                'range': f'PORCENTAGE TABLE!B{row_num}',
                'values': [[row_data.get('descuento_maximo', '')]]
            })
            batch_data.append({
                'range': f'PORCENTAGE TABLE!C{row_num}',
                'values': [[row_data.get('utilidad_minima', '')]]
            })

        batch_body = {
            'valueInputOption': 'USER_ENTERED',
            'data': batch_data
        }

        result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=batch_body
        ).execute()

        return jsonify({'message': 'Datos actualizados exitosamente', 'result': result}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ==================== RUTAS DE NAVEGACIÓN FINANZAS ====================

@app.route('/finanzas')
def finanzas():
    return render_template('finanzas.html')

@app.route('/finanzas/lineas_credito')
def finanzas_lineas_credito():
    return render_template('finanzas_lineas_credito.html')

@app.route('/finanzas/deudas_general')
def finanzas_deudas_general():
    return render_template('finanzas_deudas_general.html')

@app.route('/finanzas/ventas_problemas')
def finanzas_ventas_problemas():
    return render_template('finanzas_ventas_problemas.html')

@app.route('/finanzas/estado_cuenta_proveedores')
def finanzas_estado_cuenta_proveedores():
    return render_template('finanzas_estado_cuenta_proveedores.html')

@app.route('/finanzas/multiestados')
def finanzas_multiestados():
    return render_template('finanzas_multiestados.html')

@app.route('/finanzas/deudas_liquidadas')
def finanzas_deudas_liquidadas():
    return render_template('finanzas_deudas_liquidadas.html')

@app.route('/finanzas/calendario_deuda')
def finanzas_calendario_deuda():
    return render_template('finanzas_calendario_deuda.html')

@app.route('/finanzas/calendario_pagadas')
def finanzas_calendario_pagadas():
    return render_template('finanzas_calendario_pagadas.html')

@app.route('/finanzas/balance_semanal')
def finanzas_balance_semanal():
    return render_template('finanzas_balance_semanal.html')

@app.route('/finanzas/balance_mensual')
def finanzas_balance_mensual():
    return render_template('finanzas_balance_mensual.html')

@app.route('/finanzas/estado_lineas_credito')
def finanzas_estado_lineas_credito():
    return render_template('finanzas_estado_lineas_credito.html')

@app.route('/finanzas/presupuesto')
def finanzas_presupuesto():
    return render_template('finanzas_presupuesto.html')

@app.route('/finanzas/ventas_por_pagar')
def finanzas_ventas_por_pagar():
    return render_template('finanzas_ventas_por_pagar.html')

# ==================== RUTAS BDS FINANZAS ====================

@app.route('/bd_deudas_generales')
def bd_deudas_generales():
    return render_template('bd_deudas_generales.html')

@app.route('/bd_ventas_no_concretadas')
def bd_ventas_no_concretadas():
    return render_template('bd_ventas_no_concretadas.html')

@app.route('/bd_ordenes_compra')
def bd_ordenes_compra():
    return render_template('bd_ordenes_compra.html')

@app.route('/bd_ordenes_pagadas')
def bd_ordenes_pagadas():
    return render_template('bd_ordenes_pagadas.html')

@app.route('/bd_movimientos_financieros')
def bd_movimientos_financieros():
    return render_template('bd_movimientos_financieros.html')

# ==================== APIs FINANZAS - LECTURA DE DATOS ====================

# ID del spreadsheet de Finanzas (sincronización)
FINANZAS_SPREADSHEET_ID = '1vTydiW9EENA9i5byUkAFPsauU-y6kCHEraCgqsC75Ck'

@app.route('/api/finanzas/lineas_credito/data', methods=['GET'])
def api_finanzas_lineas_credito_data():
    """Lee datos de DIAS DE CREDITO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='DIAS DE CREDITO!A2:C'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 3:
                row.append('')
            data.append({
                'marca': row[0] if len(row) > 0 else '',
                'dias_credito': row[1] if len(row) > 1 else '',
                'credito': row[2] if len(row) > 2 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/lineas_credito/add', methods=['POST'])
def api_finanzas_lineas_credito_add():
    """Agrega un registro a DIAS DE CREDITO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='DIAS DE CREDITO!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = [
            {'range': f'DIAS DE CREDITO!A{next_row}', 'values': [[data.get('marca', '')]]},
            {'range': f'DIAS DE CREDITO!B{next_row}', 'values': [[data.get('dias_credito', '')]]},
            {'range': f'DIAS DE CREDITO!C{next_row}', 'values': [[data.get('credito', '')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Registro agregado exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/lineas_credito/update', methods=['POST'])
def api_finanzas_lineas_credito_update():
    """Actualiza datos de DIAS DE CREDITO (solo columnas B y C)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('data', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        batch_data = []
        for i, row_data in enumerate(rows):
            row_num = i + 2
            batch_data.append({'range': f'DIAS DE CREDITO!B{row_num}', 'values': [[row_data.get('dias_credito', '')]]})
            batch_data.append({'range': f'DIAS DE CREDITO!C{row_num}', 'values': [[row_data.get('credito', '')]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Datos actualizados exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/deudas_general/data', methods=['GET'])
def api_finanzas_deudas_general_data():
    """Lee datos de ESTADO DE CUENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=ECOTROS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA!A2:F'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 6:
                row.append('')
            data.append({
                'concepto': row[0] if len(row) > 0 else '',
                'nombre': row[1] if len(row) > 1 else '',
                'monto': row[2] if len(row) > 2 else '',
                'fecha_pago': row[3] if len(row) > 3 else '',
                'numero_semana': row[4] if len(row) > 4 else '',
                'task_pago': row[5] if len(row) > 5 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/ventas_problemas/data', methods=['GET'])
def api_finanzas_ventas_problemas_data():
    """Lee datos de DETECCIONES DE CANCELACIONES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='DETECCIONES DE CANCELACIONES!A1:C1'
        ).execute()
        values = result.get('values', [])
        row = values[0] if values else ['', '', '']
        while len(row) < 3:
            row.append('')
        data = {
            'totales': row[0] if len(row) > 0 else '0',
            'saldo': row[1] if len(row) > 1 else '$0',
            'cantidad': row[2] if len(row) > 2 else '0'
        }
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/estado_cuenta_proveedores/data', methods=['GET'])
def api_finanzas_estado_cuenta_proveedores_data():
    """Lee datos de ESTADO DE CUENTA PROVEEDORES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA PROVEEDORES!A2:K'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 11:
                row.append('')
            data.append({
                'marca': row[0] if len(row) > 0 else '',
                'is': row[1] if len(row) > 1 else '',
                'qty': row[2] if len(row) > 2 else '',
                'cita': row[3] if len(row) > 3 else '',
                'nombre_hoja': row[4] if len(row) > 4 else '',
                'dias_vencer': row[5] if len(row) > 5 else '',
                'fecha_cot': row[6] if len(row) > 6 else '',
                'monto_cot': row[7] if len(row) > 7 else '',
                'monto_factura': row[8] if len(row) > 8 else '',
                'dia_emision': row[9] if len(row) > 9 else '',
                'folio': row[10] if len(row) > 10 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/multiestados/data', methods=['GET'])
def api_finanzas_multiestados_data():
    """Lee datos de ESTADO DE CUENTA UNIFICADOS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA UNIFICADOS!A2:F'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 6:
                row.append('')
            data.append({
                'concepto': row[0] if len(row) > 0 else '',
                'nombre': row[1] if len(row) > 1 else '',
                'monto': row[2] if len(row) > 2 else '',
                'fecha_pago': row[3] if len(row) > 3 else '',
                'task_pago': row[4] if len(row) > 4 else '',
                'dias_restantes': row[5] if len(row) > 5 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/deudas_liquidadas/data', methods=['GET'])
def api_finanzas_deudas_liquidadas_data():
    """Lee datos de ESTADO DE CUENTA PAGADA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA PAGADA!A2:E'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 5:
                row.append('')
            data.append({
                'concepto': row[0] if len(row) > 0 else '',
                'nombre': row[1] if len(row) > 1 else '',
                'monto': row[2] if len(row) > 2 else '',
                'fecha_pago': row[3] if len(row) > 3 else '',
                'task_pago': row[4] if len(row) > 4 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/balance_semanal/data', methods=['GET'])
def api_finanzas_balance_semanal_data():
    """Lee datos de LIBERACIONES (Balance Semanal)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        # Leer semana
        semana_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='LIBERACIONES!A1'
        ).execute()
        semana = semana_result.get('values', [['']])[0][0] if semana_result.get('values') else 'SEMANA'
        # Leer datos (desde fila 3)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='LIBERACIONES!A3:D27'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 4:
                row.append('')
            data.append({
                'concepto': row[0] if len(row) > 0 else '',
                'acreditado': row[1] if len(row) > 1 else '',
                'debitado': row[2] if len(row) > 2 else '',
                'porcentaje': row[3] if len(row) > 3 else ''
            })
        # Leer total
        total_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='LIBERACIONES!A28:B28'
        ).execute()
        total_values = total_result.get('values', [[]])[0] if total_result.get('values') else []
        total = {
            'concepto': total_values[0] if len(total_values) > 0 else 'TOTAL',
            'valor': total_values[1] if len(total_values) > 1 else ''
        }
        return jsonify({'data': data, 'semana': semana, 'total': total}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/balance_mensual/data', methods=['GET'])
def api_finanzas_balance_mensual_data():
    """Lee datos de BALANCE GENERAL MENSUAL"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        # Leer mes
        mes_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='BALANCE GENERAL MENSUAL!A1'
        ).execute()
        mes = mes_result.get('values', [['']])[0][0] if mes_result.get('values') else 'MES'
        # Leer datos (desde fila 3)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='BALANCE GENERAL MENSUAL!A3:D29'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 4:
                row.append('')
            data.append({
                'concepto': row[0] if len(row) > 0 else '',
                'acreditado': row[1] if len(row) > 1 else '',
                'debitado': row[2] if len(row) > 2 else '',
                'porcentaje': row[3] if len(row) > 3 else ''
            })
        # Leer total
        total_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='BALANCE GENERAL MENSUAL!A30:B30'
        ).execute()
        total_values = total_result.get('values', [[]])[0] if total_result.get('values') else []
        total = {
            'concepto': total_values[0] if len(total_values) > 0 else 'TOTAL',
            'valor': total_values[1] if len(total_values) > 1 else ''
        }
        return jsonify({'data': data, 'mes': mes, 'total': total}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/calendario_deuda/data', methods=['GET'])
def api_finanzas_calendario_deuda_data():
    """Lee datos de DEUDA POR CONCEPTO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        meses_config = {
            'enero': 'A2:B', 'febrero': 'D2:E', 'marzo': 'G2:H', 'abril': 'J2:K',
            'mayo': 'M2:N', 'junio': 'P2:Q', 'julio': 'S2:T', 'agosto': 'V2:W',
            'septiembre': 'Y2:Z', 'octubre': 'AB2:AC', 'noviembre': 'AE2:AF', 'diciembre': 'AH2:AI'
        }
        data = {}
        for mes, rango in meses_config.items():
            result = service.spreadsheets().values().get(
                spreadsheetId=FINANZAS_SPREADSHEET_ID,
                range=f'DEUDA POR CONCEPTO!{rango}'
            ).execute()
            values = result.get('values', [])
            data[mes] = []
            for row in values:
                if len(row) >= 2:
                    data[mes].append({'concepto': row[0], 'monto': row[1]})
                elif len(row) == 1:
                    data[mes].append({'concepto': row[0], 'monto': ''})
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/calendario_pagadas/data', methods=['GET'])
def api_finanzas_calendario_pagadas_data():
    """Lee datos de DEUDA PAGADA POR CONCEPTO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        meses_config = {
            'enero': 'A2:B', 'febrero': 'D2:E', 'marzo': 'G2:H', 'abril': 'J2:K',
            'mayo': 'M2:N', 'junio': 'P2:Q', 'julio': 'S2:T', 'agosto': 'V2:W',
            'septiembre': 'Y2:Z', 'octubre': 'AB2:AC', 'noviembre': 'AE2:AF', 'diciembre': 'AH2:AI'
        }
        data = {}
        for mes, rango in meses_config.items():
            result = service.spreadsheets().values().get(
                spreadsheetId=FINANZAS_SPREADSHEET_ID,
                range=f'DEUDA PAGADA POR CONCEPTO!{rango}'
            ).execute()
            values = result.get('values', [])
            data[mes] = []
            for row in values:
                if len(row) >= 2:
                    data[mes].append({'concepto': row[0], 'monto': row[1]})
                elif len(row) == 1:
                    data[mes].append({'concepto': row[0], 'monto': ''})
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/estado_lineas_credito/data', methods=['GET'])
def api_finanzas_estado_lineas_credito_data():
    """Lee datos de PANEL DE ESTADO DE CUENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='PANEL DE ESTADO DE CUENTA!A2:F'
        ).execute()
        values = result.get('values', [])
        data = []
        for row in values:
            while len(row) < 6:
                row.append('')
            data.append({
                'marca': row[0] if len(row) > 0 else '',
                'monto_total': row[1] if len(row) > 1 else '',
                'linea_credito': row[2] if len(row) > 2 else '',
                'restante': row[3] if len(row) > 3 else '',
                'linea_usada': row[5] if len(row) > 5 else ''
            })
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/presupuesto/data', methods=['GET'])
def api_finanzas_presupuesto_data():
    """Lee datos de PRESUPUESTO SEMANAL"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='PRESUPUESTO SEMANAL!A1:B1'
        ).execute()
        values = result.get('values', [[]])[0] if result.get('values') else []
        data = {
            'label': values[0] if len(values) > 0 else 'PRESUPUESTO SEMANAL',
            'valor': values[1] if len(values) > 1 else '$0.00'
        }
        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/finanzas/ventas_por_pagar/data', methods=['GET'])
def api_finanzas_ventas_por_pagar_data():
    """Lee datos de LIBERACIONES FUTURAS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        # Leer resumen
        resumen_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='LIBERACIONES FUTURAS!A1:B3'
        ).execute()
        resumen_values = resumen_result.get('values', [])
        resumen = []
        for row in resumen_values:
            if len(row) >= 2:
                resumen.append({'label': row[0], 'valor': row[1]})
        # Leer desglose (desde fila 3)
        desglose_result = service.spreadsheets().values().get(
            spreadsheetId=FINANZAS_SPREADSHEET_ID,
            range='LIBERACIONES FUTURAS!D3:E'
        ).execute()
        desglose_values = desglose_result.get('values', [])
        desglose = []
        for row in desglose_values:
            if len(row) >= 2:
                desglose.append({'concepto': row[0], 'valor': row[1]})
        return jsonify({'data': {'resumen': resumen, 'desglose': desglose}}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ==================== APIs BDS FINANZAS - ESCRITURA ====================

# IDs de Spreadsheets para BDs de Finanzas
ECOTROS_SPREADSHEET_ID = '1ocxVz3X45i_aNOL-Slx2-u-wsF4E7-cyDJNmRTsM5PU'
VENTASRETORNADAS_SPREADSHEET_ID = '1qIWn_UuOpsF4VO5QOZlDjXneZtdP1jy9PAeScJ-5O1A'
PORLIQUIDAR_SPREADSHEET_ID = '1soUT-OcdAZX-aStHmePQatz94TjjDH1lsahs99T7pTI'
BALANCESSEMANALES_SPREADSHEET_ID = '1YEpxA-AEjgkxl2ZsTitV9fns7aJgin4vVtF72hIjyj8'
BALANCESMENSUALES_SPREADSHEET_ID = '1EQmljLG5U-SZQgKKm_3qX4ujs3PO9kGR_4lpW71DULg'
ESTADO_PRECIOS_SPREADSHEET_ID = '1hnRkPS2LlWXTcFyBaTqgmP7WGV8Fk-k9nLAFkNqP_kU'

@app.route('/api/bd_deudas_generales/add', methods=['POST'])
def api_bd_deudas_generales_add():
    """Agrega un registro a ECOTROS - ESTADO DE CUENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=ECOTROS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = [
            {'range': f'ESTADO DE CUENTA!A{next_row}', 'values': [[data.get('concepto', '')]]},
            {'range': f'ESTADO DE CUENTA!B{next_row}', 'values': [[data.get('nombre', '')]]},
            {'range': f'ESTADO DE CUENTA!C{next_row}', 'values': [[data.get('monto', '')]]},
            {'range': f'ESTADO DE CUENTA!D{next_row}', 'values': [[data.get('fecha_pago', '')]]},
            {'range': f'ESTADO DE CUENTA!E{next_row}', 'values': [[data.get('numero_semana', '')]]},
            {'range': f'ESTADO DE CUENTA!F{next_row}', 'values': [[data.get('pagado', 'FALSE')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=ECOTROS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Registro agregado exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_deudas_generales/bulk', methods=['POST'])
def api_bd_deudas_generales_bulk():
    """Carga masiva a ECOTROS - ESTADO DE CUENTA"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('rows', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=ECOTROS_SPREADSHEET_ID,
            range='ESTADO DE CUENTA!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = []
        for i, row_data in enumerate(rows):
            current_row = next_row + i
            batch_data.append({'range': f'ESTADO DE CUENTA!A{current_row}', 'values': [[row_data.get('concepto', '')]]})
            batch_data.append({'range': f'ESTADO DE CUENTA!B{current_row}', 'values': [[row_data.get('nombre', '')]]})
            batch_data.append({'range': f'ESTADO DE CUENTA!C{current_row}', 'values': [[row_data.get('monto', '')]]})
            batch_data.append({'range': f'ESTADO DE CUENTA!D{current_row}', 'values': [[row_data.get('fecha_pago', '')]]})
            batch_data.append({'range': f'ESTADO DE CUENTA!E{current_row}', 'values': [[row_data.get('numero_semana', '')]]})
            batch_data.append({'range': f'ESTADO DE CUENTA!F{current_row}', 'values': [[row_data.get('pagado', 'FALSE')]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=ECOTROS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': f'{len(rows)} registros agregados exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ventas_no_concretadas/add', methods=['POST'])
def api_bd_ventas_no_concretadas_add():
    """Agrega un registro a VENTASRETORNADAS - BD VENTAS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=VENTASRETORNADAS_SPREADSHEET_ID,
            range='BD VENTAS!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = max(len(existing) + 1, 6)
        batch_data = [
            {'range': f'BD VENTAS!A{next_row}', 'values': [[data.get('numero_id', '')]]},
            {'range': f'BD VENTAS!C{next_row}', 'values': [[data.get('estado', '')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=VENTASRETORNADAS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Registro agregado exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ventas_no_concretadas/bulk', methods=['POST'])
def api_bd_ventas_no_concretadas_bulk():
    """Carga masiva a VENTASRETORNADAS - BD VENTAS"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('rows', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=VENTASRETORNADAS_SPREADSHEET_ID,
            range='BD VENTAS!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = max(len(existing) + 1, 6)
        batch_data = []
        for i, row_data in enumerate(rows):
            current_row = next_row + i
            batch_data.append({'range': f'BD VENTAS!A{current_row}', 'values': [[row_data.get('numero_id', '')]]})
            batch_data.append({'range': f'BD VENTAS!C{current_row}', 'values': [[row_data.get('estado', '')]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=VENTASRETORNADAS_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': f'{len(rows)} registros agregados exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ordenes_compra/add', methods=['POST'])
def api_bd_ordenes_compra_add():
    """Agrega un registro a PORLIQUIDAR - BD MO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            range='BD MO!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = [
            {'range': f'BD MO!A{next_row}', 'values': [[data.get('numero_id', '')]]},
            {'range': f'BD MO!E{next_row}', 'values': [[data.get('estatus', '')]]},
            {'range': f'BD MO!I{next_row}', 'values': [[data.get('saldo_cotizacion', '')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Orden registrada exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ordenes_compra/bulk', methods=['POST'])
def api_bd_ordenes_compra_bulk():
    """Carga masiva a PORLIQUIDAR - BD MO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('rows', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            range='BD MO!A:A'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = []
        for i, row_data in enumerate(rows):
            current_row = next_row + i
            batch_data.append({'range': f'BD MO!A{current_row}', 'values': [[row_data.get('numero_id', '')]]})
            batch_data.append({'range': f'BD MO!E{current_row}', 'values': [[row_data.get('estatus', '')]]})
            batch_data.append({'range': f'BD MO!I{current_row}', 'values': [[row_data.get('saldo_cotizacion', '')]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': f'{len(rows)} ordenes registradas exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ordenes_pagadas/add', methods=['POST'])
def api_bd_ordenes_pagadas_add():
    """Agrega un registro a PORLIQUIDAR - BD LIBERACIONES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            range='BD LIBERACIONES!C:C'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = [
            {'range': f'BD LIBERACIONES!C{next_row}', 'values': [[data.get('numero_id', '')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': 'Orden pagada registrada exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_ordenes_pagadas/bulk', methods=['POST'])
def api_bd_ordenes_pagadas_bulk():
    """Carga masiva a PORLIQUIDAR - BD LIBERACIONES"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('rows', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            range='BD LIBERACIONES!C:C'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1
        batch_data = []
        for i, row_data in enumerate(rows):
            current_row = next_row + i
            batch_data.append({'range': f'BD LIBERACIONES!C{current_row}', 'values': [[row_data.get('numero_id', '')]]})
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=PORLIQUIDAR_SPREADSHEET_ID,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': f'{len(rows)} ordenes pagadas registradas exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_movimientos_financieros/add', methods=['POST'])
def api_bd_movimientos_financieros_add():
    """Agrega un registro a BALANCES SEMANALES o MENSUALES segun la fecha"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        from googleapiclient.discovery import build
        from datetime import datetime, timedelta
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Determinar si es semanal o mensual
        fecha_str = data.get('fecha_liberacion', '')
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d') if fecha_str else datetime.now()
        today = datetime.now()

        # Calcular inicio de semana actual y pasada
        start_of_week = today - timedelta(days=today.weekday())
        start_of_last_week = start_of_week - timedelta(days=7)

        # Calcular inicio del mes pasado
        start_of_month = today.replace(day=1)
        start_of_last_month = (start_of_month - timedelta(days=1)).replace(day=1)
        end_of_last_month = start_of_month - timedelta(days=1)

        # Decidir destino
        if fecha >= start_of_last_week:
            # Semana actual o pasada -> BALANCESSEMANALES
            spreadsheet_id = BALANCESSEMANALES_SPREADSHEET_ID
            destino = 'semanal'
        elif start_of_last_month <= fecha <= end_of_last_month:
            # Mes pasado -> BALANCESMENSUALES
            spreadsheet_id = BALANCESMENSUALES_SPREADSHEET_ID
            destino = 'mensual'
        else:
            # Default: semanal
            spreadsheet_id = BALANCESSEMANALES_SPREADSHEET_ID
            destino = 'semanal'

        # Buscar ultima fila con valor en columna E
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range='BD!E:E'
        ).execute()
        existing = result.get('values', [])
        next_row = len(existing) + 1

        # Obtener metadata de la hoja para verificar limite de filas
        sheet_metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        bd_sheet = None
        for sheet in sheet_metadata.get('sheets', []):
            if sheet.get('properties', {}).get('title') == 'BD':
                bd_sheet = sheet
                break

        if bd_sheet:
            current_max_rows = bd_sheet.get('properties', {}).get('gridProperties', {}).get('rowCount', 1000)
            # Si la siguiente fila excede el limite, expandir la hoja
            if next_row >= current_max_rows:
                rows_to_add = 1000  # Agregar 1000 filas mas
                sheet_id = bd_sheet.get('properties', {}).get('sheetId')
                service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        'requests': [{
                            'appendDimension': {
                                'sheetId': sheet_id,
                                'dimension': 'ROWS',
                                'length': rows_to_add
                            }
                        }]
                    }
                ).execute()

        batch_data = [
            {'range': f'BD!A{next_row}', 'values': [[data.get('fecha_liberacion', '')]]},
            {'range': f'BD!C{next_row}', 'values': [[data.get('numero_id', '')]]},
            {'range': f'BD!E{next_row}', 'values': [[data.get('descripcion', '')]]},
            {'range': f'BD!F{next_row}', 'values': [[data.get('monto_acreditado', '')]]},
            {'range': f'BD!G{next_row}', 'values': [[data.get('monto_debitado', '')]]}
        ]
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
        ).execute()
        return jsonify({'message': f'Movimiento registrado en balance {destino}'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/bd_movimientos_financieros/bulk', methods=['POST'])
def api_bd_movimientos_financieros_bulk():
    """Carga masiva a BALANCES SEMANALES o MENSUALES segun las fechas"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('rows', [])
        from googleapiclient.discovery import build
        from datetime import datetime, timedelta
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        start_of_last_week = start_of_week - timedelta(days=7)
        start_of_month = today.replace(day=1)
        start_of_last_month = (start_of_month - timedelta(days=1)).replace(day=1)
        end_of_last_month = start_of_month - timedelta(days=1)

        semanales = []
        mensuales = []

        for row_data in rows:
            fecha_str = row_data.get('fecha_liberacion', '')
            try:
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
            except:
                fecha = today

            if fecha >= start_of_last_week:
                semanales.append(row_data)
            elif start_of_last_month <= fecha <= end_of_last_month:
                mensuales.append(row_data)
            else:
                semanales.append(row_data)

        # Procesar semanales
        if semanales:
            result = service.spreadsheets().values().get(
                spreadsheetId=BALANCESSEMANALES_SPREADSHEET_ID,
                range='BD!E:E'
            ).execute()
            existing = result.get('values', [])
            next_row = len(existing) + 1
            last_row_needed = next_row + len(semanales) - 1

            # Verificar y expandir hoja si es necesario
            sheet_metadata = service.spreadsheets().get(spreadsheetId=BALANCESSEMANALES_SPREADSHEET_ID).execute()
            for sheet in sheet_metadata.get('sheets', []):
                if sheet.get('properties', {}).get('title') == 'BD':
                    current_max_rows = sheet.get('properties', {}).get('gridProperties', {}).get('rowCount', 1000)
                    if last_row_needed >= current_max_rows:
                        rows_to_add = last_row_needed - current_max_rows + 1000
                        sheet_id = sheet.get('properties', {}).get('sheetId')
                        service.spreadsheets().batchUpdate(
                            spreadsheetId=BALANCESSEMANALES_SPREADSHEET_ID,
                            body={'requests': [{'appendDimension': {'sheetId': sheet_id, 'dimension': 'ROWS', 'length': rows_to_add}}]}
                        ).execute()
                    break

            batch_data = []
            for i, row_data in enumerate(semanales):
                current_row = next_row + i
                batch_data.append({'range': f'BD!A{current_row}', 'values': [[row_data.get('fecha_liberacion', '')]]})
                batch_data.append({'range': f'BD!C{current_row}', 'values': [[row_data.get('numero_id', '')]]})
                batch_data.append({'range': f'BD!E{current_row}', 'values': [[row_data.get('descripcion', '')]]})
                batch_data.append({'range': f'BD!F{current_row}', 'values': [[row_data.get('monto_acreditado', '')]]})
                batch_data.append({'range': f'BD!G{current_row}', 'values': [[row_data.get('monto_debitado', '')]]})
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=BALANCESSEMANALES_SPREADSHEET_ID,
                body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
            ).execute()

        # Procesar mensuales
        if mensuales:
            result = service.spreadsheets().values().get(
                spreadsheetId=BALANCESMENSUALES_SPREADSHEET_ID,
                range='BD!E:E'
            ).execute()
            existing = result.get('values', [])
            next_row = len(existing) + 1
            last_row_needed = next_row + len(mensuales) - 1

            # Verificar y expandir hoja si es necesario
            sheet_metadata = service.spreadsheets().get(spreadsheetId=BALANCESMENSUALES_SPREADSHEET_ID).execute()
            for sheet in sheet_metadata.get('sheets', []):
                if sheet.get('properties', {}).get('title') == 'BD':
                    current_max_rows = sheet.get('properties', {}).get('gridProperties', {}).get('rowCount', 1000)
                    if last_row_needed >= current_max_rows:
                        rows_to_add = last_row_needed - current_max_rows + 1000
                        sheet_id = sheet.get('properties', {}).get('sheetId')
                        service.spreadsheets().batchUpdate(
                            spreadsheetId=BALANCESMENSUALES_SPREADSHEET_ID,
                            body={'requests': [{'appendDimension': {'sheetId': sheet_id, 'dimension': 'ROWS', 'length': rows_to_add}}]}
                        ).execute()
                    break

            batch_data = []
            for i, row_data in enumerate(mensuales):
                current_row = next_row + i
                batch_data.append({'range': f'BD!A{current_row}', 'values': [[row_data.get('fecha_liberacion', '')]]})
                batch_data.append({'range': f'BD!C{current_row}', 'values': [[row_data.get('numero_id', '')]]})
                batch_data.append({'range': f'BD!E{current_row}', 'values': [[row_data.get('descripcion', '')]]})
                batch_data.append({'range': f'BD!F{current_row}', 'values': [[row_data.get('monto_acreditado', '')]]})
                batch_data.append({'range': f'BD!G{current_row}', 'values': [[row_data.get('monto_debitado', '')]]})
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=BALANCESMENSUALES_SPREADSHEET_ID,
                body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
            ).execute()

        return jsonify({
            'message': 'Carga completada',
            'semanal': len(semanales),
            'mensual': len(mensuales)
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# --- ENDPOINTS PARA ESTADO DE PRECIOS ---
@app.route('/estado_precios')
def estado_precios():
    return render_template('estado_precios.html')

@app.route('/api/estado_precios/data', methods=['GET'])
def api_estado_precios_data():
    """Lee datos de LISTA L - Estado de Precios"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Leer todas las columnas necesarias: A, B, C, D, F, G, H, K, L, M, N
        result = service.spreadsheets().values().batchGet(
            spreadsheetId=ESTADO_PRECIOS_SPREADSHEET_ID,
            ranges=[
                'LISTA L!A2:A',  # CATEGORIA
                'LISTA L!B2:B',  # ID (editable)
                'LISTA L!C2:C',  # MARCA
                'LISTA L!D2:D',  # SKU
                'LISTA L!F2:F',  # CLASIFICACION
                'LISTA L!G2:G',  # ACTUAL PRICE
                'LISTA L!H2:H',  # NEW PRECIO
                'LISTA L!K2:K',  # DESCUENTO O INCREMENTO (editable)
                'LISTA L!L2:L',  # UTILIDAD
                'LISTA L!M2:M',  # UTILIDAD %
                'LISTA L!N2:N',  # RANGO
            ]
        ).execute()

        value_ranges = result.get('valueRanges', [])

        # Extraer valores de cada columna
        col_categoria = [row[0] if row else '' for row in value_ranges[0].get('values', [])]
        col_id = [row[0] if row else '' for row in value_ranges[1].get('values', [])]
        col_marca = [row[0] if row else '' for row in value_ranges[2].get('values', [])]
        col_sku = [row[0] if row else '' for row in value_ranges[3].get('values', [])]
        col_clasificacion = [row[0] if row else '' for row in value_ranges[4].get('values', [])]
        col_actual_price = [row[0] if row else '' for row in value_ranges[5].get('values', [])]
        col_new_precio = [row[0] if row else '' for row in value_ranges[6].get('values', [])]
        col_descuento = [row[0] if row else '' for row in value_ranges[7].get('values', [])]
        col_utilidad = [row[0] if row else '' for row in value_ranges[8].get('values', [])]
        col_utilidad_pct = [row[0] if row else '' for row in value_ranges[9].get('values', [])]
        col_rango = [row[0] if row else '' for row in value_ranges[10].get('values', [])]

        # Encontrar la longitud máxima
        max_len = max(len(col_categoria), len(col_id), len(col_marca), len(col_sku),
                      len(col_clasificacion), len(col_actual_price), len(col_new_precio),
                      len(col_descuento), len(col_utilidad), len(col_utilidad_pct), len(col_rango))

        # Construir datos (solo filas donde ID no esté vacío)
        data = []
        for i in range(max_len):
            id_value = col_id[i] if i < len(col_id) else ''
            # Solo incluir filas donde el ID no esté vacío
            if id_value and id_value.strip():
                data.append({
                    'row_index': i + 2,
                    'categoria': col_categoria[i] if i < len(col_categoria) else '',
                    'id': id_value,
                    'marca': col_marca[i] if i < len(col_marca) else '',
                    'sku': col_sku[i] if i < len(col_sku) else '',
                    'clasificacion': col_clasificacion[i] if i < len(col_clasificacion) else '',
                    'actual_price': col_actual_price[i] if i < len(col_actual_price) else '',
                    'new_precio': col_new_precio[i] if i < len(col_new_precio) else '',
                    'descuento_incremento': col_descuento[i] if i < len(col_descuento) else '',
                    'utilidad': col_utilidad[i] if i < len(col_utilidad) else '',
                    'utilidad_pct': col_utilidad_pct[i] if i < len(col_utilidad_pct) else '',
                    'rango': col_rango[i] if i < len(col_rango) else ''
                })

        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/estado_precios/update', methods=['POST'])
def api_estado_precios_update():
    """Actualiza datos de LISTA L - Solo columnas B (ID) y K (DESCUENTO O INCREMENTO)"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        data = request.get_json()
        rows = data.get('data', [])
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        batch_data = []
        for row_data in rows:
            row_num = row_data.get('row_index', 0)
            if row_num > 0:
                # Agregar % al descuento si tiene valor y no termina en %
                descuento_val = row_data.get('descuento_incremento', '')
                if descuento_val and not str(descuento_val).endswith('%'):
                    descuento_val = str(descuento_val) + '%'
                # Solo actualizar columnas B (ID) y K (DESCUENTO O INCREMENTO)
                batch_data.append({'range': f'LISTA L!B{row_num}', 'values': [[row_data.get('id', '')]]})
                batch_data.append({'range': f'LISTA L!K{row_num}', 'values': [[descuento_val]]})

        if batch_data:
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=ESTADO_PRECIOS_SPREADSHEET_ID,
                body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
            ).execute()

        return jsonify({'message': 'Datos actualizados exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/estado_precios/bulk', methods=['POST'])
def api_estado_precios_bulk():
    """Carga masiva CSV - Solo columnas ID y DESCUENTO O INCREMENTO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        import io, csv

        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        filename = file.filename.lower()

        if not filename.endswith('.csv'):
            return jsonify({'error': 'Solo se permiten archivos CSV'}), 400

        # Leer CSV
        try:
            stream = io.StringIO(file.stream.read().decode('utf-8'))
        except Exception:
            stream = io.StringIO(file.stream.read().decode('latin-1'))

        reader = csv.DictReader(stream)

        if reader.fieldnames is None:
            return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400

        # Verificar columnas requeridas (ID y DESCUENTO O INCREMENTO)
        required_cols = ['ID', 'DESCUENTO O INCREMENTO']
        missing = [col for col in required_cols if col not in reader.fieldnames]
        if missing:
            return jsonify({'error': f'Faltan columnas requeridas: {missing}. El CSV debe tener: ID, DESCUENTO O INCREMENTO'}), 400

        rows = list(reader)

        if not rows:
            return jsonify({'error': 'El archivo CSV está vacío'}), 400

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Preparar datos para actualizar
        batch_data = []
        for i, row in enumerate(rows):
            row_num = i + 2  # Empezar desde fila 2
            # Agregar % al descuento si tiene valor y no termina en %
            descuento_val = row.get('DESCUENTO O INCREMENTO', '')
            if descuento_val and not str(descuento_val).endswith('%'):
                descuento_val = str(descuento_val) + '%'
            batch_data.append({'range': f'LISTA L!B{row_num}', 'values': [[row.get('ID', '')]]})
            batch_data.append({'range': f'LISTA L!K{row_num}', 'values': [[descuento_val]]})

        if batch_data:
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=ESTADO_PRECIOS_SPREADSHEET_ID,
                body={'valueInputOption': 'USER_ENTERED', 'data': batch_data}
            ).execute()

        return jsonify({'message': f'Se actualizaron {len(rows)} registros exitosamente'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/estado_precios/errores', methods=['GET'])
def api_estado_precios_errores():
    """Lee datos de ID ERRORES - Errores Estado de Precios"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Leer columnas: A (ID), B (SKU), C (ERROR DETECTADO), D (SOLUCION MARCA), E (SOLUCION COSTO), F (SOLUCION ENVIO)
        result = service.spreadsheets().values().batchGet(
            spreadsheetId=ESTADO_PRECIOS_SPREADSHEET_ID,
            ranges=[
                'ID ERRORES!A2:A',  # ID
                'ID ERRORES!B2:B',  # SKU
                'ID ERRORES!C2:C',  # ERROR DETECTADO
                'ID ERRORES!D2:D',  # SOLUCION PARA MARCA
                'ID ERRORES!E2:E',  # SOLUCION PARA COSTO
                'ID ERRORES!F2:F',  # SOLUCION PARA ENVIO
            ]
        ).execute()

        value_ranges = result.get('valueRanges', [])

        # Extraer valores de cada columna
        col_id = [row[0] if row else '' for row in value_ranges[0].get('values', [])]
        col_sku = [row[0] if row else '' for row in value_ranges[1].get('values', [])]
        col_error = [row[0] if row else '' for row in value_ranges[2].get('values', [])]
        col_sol_marca = [row[0] if row else '' for row in value_ranges[3].get('values', [])]
        col_sol_costo = [row[0] if row else '' for row in value_ranges[4].get('values', [])]
        col_sol_envio = [row[0] if row else '' for row in value_ranges[5].get('values', [])]

        # Encontrar la longitud máxima
        max_len = max(len(col_id), len(col_sku), len(col_error),
                      len(col_sol_marca), len(col_sol_costo), len(col_sol_envio), 1) if any([col_id, col_sku, col_error]) else 0

        # Construir datos (solo filas donde hay algún dato)
        data = []
        for i in range(max_len):
            id_val = col_id[i] if i < len(col_id) else ''
            sku_val = col_sku[i] if i < len(col_sku) else ''
            error_val = col_error[i] if i < len(col_error) else ''
            # Solo incluir si hay al menos ID o SKU o ERROR
            if id_val or sku_val or error_val:
                data.append({
                    'row_index': i + 2,
                    'id': id_val,
                    'sku': sku_val,
                    'error_detectado': error_val,
                    'solucion_marca': col_sol_marca[i] if i < len(col_sol_marca) else '',
                    'solucion_costo': col_sol_costo[i] if i < len(col_sol_costo) else '',
                    'solucion_envio': col_sol_envio[i] if i < len(col_sol_envio) else ''
                })

        return jsonify({'data': data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# --- ENDPOINTS PARA PRODUCTOS OLVIDADOS ---

# Configuracion por defecto de IMPULSO
IMPULSO_DEFAULT_URL = 'https://docs.google.com/spreadsheets/d/1LDZasT6HhRheM-IYWbw_tZLNZ5cDZgwNamk84UsIdIY/edit?gid=2034740744#gid=2034740744'
IMPULSO_DEFAULT_HOJA = {
    'HOJA': 'IMPULSE ORDER',
    'MARCA': 'A1',
    'PORCENTAJE': 'E1',
    'TABLA': 'A3:F',
    'TRUPER': 'C1',
    '0.1': 'F1'
}

def _init_impulso_config(uid):
    """Auto-crea la configuracion de IMPULSO en Firebase si no existe"""
    area_ref = db.collection('Areas').document(uid)
    area_doc = area_ref.get()
    if not area_doc.exists:
        area_ref.set({'IMPULSO': IMPULSO_DEFAULT_URL})
    else:
        area_data = area_doc.to_dict()
        if not area_data.get('IMPULSO'):
            area_ref.update({'IMPULSO': IMPULSO_DEFAULT_URL})
    hoja_ref = area_ref.collection('Hojas').document('impulse')
    hoja_doc = hoja_ref.get()
    if not hoja_doc.exists:
        hoja_ref.set(IMPULSO_DEFAULT_HOJA)

@app.route('/productos_olvidados')
def productos_olvidados():
    return render_template('productos_olvidados.html')

@app.route('/api/productos_olvidados/data', methods=['GET'])
def api_productos_olvidados_data():
    """Lee datos de IMPULSO - Productos Olvidados"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        _init_impulso_config(uid)

        # Obtener configuracion de Firebase
        area_doc = db.collection('Areas').document(uid).get()
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('IMPULSO')

        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet invalida'}), 400
        spreadsheet_id = match.group(1)

        # Obtener configuracion de hojas
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('impulse')
        hoja_doc = hoja_ref.get()
        hoja_data = hoja_doc.to_dict()

        nombre_hoja = hoja_data.get('HOJA', 'IMPULSE ORDER')
        celda_marca = hoja_data.get('MARCA', 'A1')
        celda_porcentaje = hoja_data.get('PORCENTAJE', 'E1')
        rango_tabla = hoja_data.get('TABLA', 'A3:F')
        celda_truper = hoja_data.get('TRUPER', 'C1')
        celda_valor = hoja_data.get('0.1', 'F1')

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # Leer todos los datos necesarios en batch
        ranges = [
            f"{nombre_hoja}!{celda_marca}",
            f"{nombre_hoja}!{celda_porcentaje}",
            f"{nombre_hoja}!{rango_tabla}",
            f"{nombre_hoja}!{celda_truper}",
            f"{nombre_hoja}!{celda_valor}",
        ]

        result = service.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=ranges
        ).execute()

        value_ranges = result.get('valueRanges', [])

        marca = ''
        if len(value_ranges) > 0 and value_ranges[0].get('values'):
            marca = value_ranges[0]['values'][0][0] if value_ranges[0]['values'][0] else ''

        porcentaje = ''
        if len(value_ranges) > 1 and value_ranges[1].get('values'):
            porcentaje = value_ranges[1]['values'][0][0] if value_ranges[1]['values'][0] else ''

        tabla = []
        tabla_headers = []
        if len(value_ranges) > 2 and value_ranges[2].get('values'):
            all_rows = value_ranges[2]['values']
            if len(all_rows) > 0:
                tabla_headers = all_rows[0]
                tabla = all_rows[1:] if len(all_rows) > 1 else []

        truper_actual = ''
        if len(value_ranges) > 3 and value_ranges[3].get('values'):
            truper_actual = value_ranges[3]['values'][0][0] if value_ranges[3]['values'][0] else ''

        valor_f1 = ''
        if len(value_ranges) > 4 and value_ranges[4].get('values'):
            valor_f1 = value_ranges[4]['values'][0][0] if value_ranges[4]['values'][0] else ''

        # Obtener opciones de validacion de datos para celda C1 (TRUPER)
        truper_opciones = []
        try:
            # Obtener la fila y columna de celda_truper
            sheet_meta = service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                ranges=[f"{nombre_hoja}!{celda_truper}"],
                fields='sheets.data.rowData.values.dataValidation'
            ).execute()

            sheets = sheet_meta.get('sheets', [])
            if sheets:
                data_list = sheets[0].get('data', [])
                if data_list:
                    row_data = data_list[0].get('rowData', [])
                    if row_data:
                        values = row_data[0].get('values', [])
                        if values:
                            dv = values[0].get('dataValidation', {})
                            condition = dv.get('condition', {})
                            if condition.get('type') == 'ONE_OF_LIST':
                                cond_values = condition.get('values', [])
                                truper_opciones = [v.get('userEnteredValue', '') for v in cond_values if v.get('userEnteredValue')]
                            elif condition.get('type') == 'ONE_OF_RANGE':
                                # Si la validacion es un rango, leer ese rango
                                cond_values = condition.get('values', [])
                                if cond_values:
                                    ref_range = cond_values[0].get('userEnteredValue', '')
                                    if ref_range.startswith('='):
                                        ref_range = ref_range[1:]
                                    ref_result = service.spreadsheets().values().get(
                                        spreadsheetId=spreadsheet_id,
                                        range=ref_range
                                    ).execute()
                                    ref_values = ref_result.get('values', [])
                                    truper_opciones = [row[0] for row in ref_values if row]
        except Exception:
            pass

        return jsonify({
            'marca': marca,
            'porcentaje': porcentaje,
            'tabla': tabla,
            'tabla_headers': tabla_headers,
            'truper_actual': truper_actual,
            'truper_opciones': truper_opciones,
            'valor_f1': valor_f1
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/productos_olvidados/update', methods=['POST'])
def api_productos_olvidados_update():
    """Actualiza valores editables en IMPULSO"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        data = request.get_json()
        _init_impulso_config(uid)

        area_doc = db.collection('Areas').document(uid).get()
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('IMPULSO')

        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            return jsonify({'error': 'URL de spreadsheet invalida'}), 400
        spreadsheet_id = match.group(1)

        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('impulse')
        hoja_doc = hoja_ref.get()
        hoja_data = hoja_doc.to_dict()

        nombre_hoja = hoja_data.get('HOJA', 'IMPULSE ORDER')
        celda_truper = hoja_data.get('TRUPER', 'C1')
        celda_valor = hoja_data.get('0.1', 'F1')

        from googleapiclient.discovery import build
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)

        batch_data = []

        if 'truper' in data:
            batch_data.append({
                'range': f"{nombre_hoja}!{celda_truper}",
                'values': [[data['truper']]]
            })

        if 'valor_f1' in data:
            batch_data.append({
                'range': f"{nombre_hoja}!{celda_valor}",
                'values': [[data['valor_f1']]]
            })

        if batch_data:
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    'valueInputOption': 'USER_ENTERED',
                    'data': batch_data
                }
            ).execute()

        return jsonify({'status': 'ok'}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5001))
    print(f'Servidor Flask corriendo en puerto {port}...')
    app.run(host='0.0.0.0', port=port, debug=False)