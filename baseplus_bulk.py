from flask import request, jsonify
from werkzeug.utils import secure_filename
import os
import csv
import io
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from firebase_admin import auth
from google.cloud import firestore
from flask import current_app as app

@app.route('/api/baseplus_bulk', methods=['POST'])
def baseplus_bulk():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'No token provided'}), 401
    id_token = auth_header.split(' ')[1]
    import traceback
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        import sys
        def log(msg, *args):
            print('[BULK]', msg, *args)
            sys.stdout.flush()
        log('UID:', uid)
        db = firestore.Client.from_service_account_json('supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json')
        hoja_ref = db.collection('Areas').document(uid).collection('Hojas').document('BASEPLUS')
        hoja_doc = hoja_ref.get()
        if not hoja_doc.exists:
            print('[BULK] No se encontró la configuración de hoja/rango')
            return jsonify({'error': 'No se encontró la configuración de hoja/rango'}), 404
        hoja_data = hoja_doc.to_dict()
        nombre_hoja = hoja_data.get('Hoja', 'Sheet1')
        ubicaciones = {k: v for k, v in hoja_data.items() if k != 'Hoja'}
        log('Ubicaciones:', ubicaciones)
        area_doc = db.collection('Areas').document(uid).get()
        if not area_doc.exists:
            print('[BULK] No se encontró el área para este usuario')
            return jsonify({'error': 'No se encontró el área para este usuario'}), 404
        area_data = area_doc.to_dict()
        spreadsheet_url = area_data.get('BASEPLUS')
        import re
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", spreadsheet_url)
        if not match:
            print('[BULK] URL de spreadsheet inválida:', spreadsheet_url)
            return jsonify({'error': 'URL de spreadsheet inválida'}), 400
        spreadsheet_id = match.group(1)
        if 'file' not in request.files:
            print('[BULK] No file uploaded')
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = file.filename.lower()
        campos_config = list(ubicaciones.keys())
        log('Campos config:', campos_config)
        rows = []
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode('utf-8'))
            reader = csv.DictReader(stream)
            log('CSV headers:', reader.fieldnames)
            if reader.fieldnames is None:
                log('El archivo CSV no tiene encabezados.')
                return jsonify({'error': 'El archivo CSV no tiene encabezados.'}), 400
            missing = [campo for campo in campos_config if campo not in reader.fieldnames]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            rows = list(reader)
            log('Primeras filas:', rows[:3])
        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            log('XLSX headers:', headers)
            missing = [campo for campo in campos_config if campo not in headers]
            if missing:
                log('Faltan columnas requeridas:', missing)
                return jsonify({'error': f'Faltan columnas requeridas: {missing}'}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            log('Primeras filas:', rows[:3])
        else:
            log('Formato de archivo no permitido:', filename)
            return jsonify({'error': 'Solo se permiten archivos CSV o XLSX'}), 400
        rows_preview = rows[:100]
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = 'supratechweb-firebase-adminsdk-fbsvc-8d4aa68a75.json'
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        sheet = service.spreadsheets()
        for campo, rango in ubicaciones.items():
            rango_col = f"{nombre_hoja}!{rango}"
            log(f'Limpiando rango {rango_col}')
            sheet.values().clear(
                spreadsheetId=spreadsheet_id,
                range=rango_col,
                body={}
            ).execute()
        for campo, rango in ubicaciones.items():
            col = rango.split(':')[0]
            fila = int(''.join(filter(str.isdigit, rango)))
            values = [[row.get(campo, '')] for row in rows]
            log(f'Escribiendo campo {campo} en {col}{fila}:{col}{fila+len(values)-1} con valores:', values[:3])
            rango_celda = f"{nombre_hoja}!{col}{fila}:{col}{fila+len(values)-1}"
            sheet.values().update(
                spreadsheetId=spreadsheet_id,
                range=rango_celda,
                valueInputOption='USER_ENTERED',
                body={'values': values}
            ).execute()
        log('Escritura masiva terminada')
        return jsonify({'status': 'ok', 'rows': len(rows), 'preview': rows_preview})
    except Exception as e:
        log('ERROR:', str(e))
        log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
